# -*- coding: utf-8 -*-
"""
邮箱收件箱阅读器 - 类 Foxmail 简易版
通过邮箱和密码直接登录 IMAP，读取最近 5 封邮件；
支持保存/切换已登录账号，从邮件正文提取验证码。
"""

from __future__ import annotations

import imaplib
import email
import re
import json
import os
import socket
import sys
import subprocess
import threading
import datetime as _dt
import shlex
import urllib.request
import urllib.error
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
from email.header import decode_header

import requests

def _style_lookup(style: ttk.Style, stylename: str, option: str, default: str) -> str:
    """ttk.Style.lookup 的容错封装（返回 default 而非空字符串）。"""
    try:
        v = style.lookup(stylename, option)
        return v if v else default
    except Exception:
        return default


class RoundedBox(tk.Canvas):
    """
    一个自绘圆角容器，用来包裹输入控件，实现“圆角输入框”效果。
    用法：box = RoundedBox(parent); box.grid(...); 在 box.inner 内部放 Entry/Combobox 等。
    """

    def __init__(
        self,
        master,
        *,
        radius: int = 10,
        # 让内容更贴近边框（用户偏好“几乎无距离”）
        pad_x: int = 4,
        pad_y: int = 2,
        height: int = 32,
        border: str = "#d0d0d0",
        focus_border: str = "#3b82f6",
        fill: str = "#ffffff",
        outer_bg: str | None = None,
        **kwargs,
    ):
        self.radius = radius
        self.pad_x = pad_x
        self.pad_y = pad_y
        self._border = border
        self._focus_border = focus_border
        self._fill = fill
        self._focused = False

        if outer_bg is None:
            # 有些 ttk 容器（例如 LabelFrame）不支持 cget("bg")，会抛 TclError: unknown option "-bg"
            # 这里尽量稳妥：优先走 style.lookup，其次再兜底为通用颜色。
            try:
                style = ttk.Style()
                outer_bg = _style_lookup(style, "TFrame", "background", "")
            except Exception:
                outer_bg = ""
            if not outer_bg:
                outer_bg = "#f0f0f0"

        super().__init__(
            master,
            highlightthickness=0,
            bd=0,
            bg=outer_bg,
            height=height,
            **kwargs,
        )
        self.inner = tk.Frame(self, bg=self._fill)
        self._win = self.create_window((self.pad_x, self.pad_y), window=self.inner, anchor="nw")
        self.bind("<Configure>", self._on_configure)

    def set_focused(self, focused: bool) -> None:
        self._focused = bool(focused)
        self._redraw()

    def _on_configure(self, _event=None) -> None:
        self._redraw()

    def _redraw(self) -> None:
        w = max(int(self.winfo_width()), 2)
        h = max(int(self.winfo_height()), 2)
        r = min(int(self.radius), max(1, min(w, h) // 2))

        self.delete("round")
        border = self._focus_border if self._focused else self._border

        # 圆角矩形（smooth polygon）
        pts = [
            r,
            0,
            w - r,
            0,
            w,
            0,
            w,
            r,
            w,
            h - r,
            w,
            h,
            w - r,
            h,
            r,
            h,
            0,
            h,
            0,
            h - r,
            0,
            r,
            0,
            0,
        ]
        self.create_polygon(
            pts,
            smooth=True,
            splinesteps=20,
            fill=self._fill,
            outline=border,
            width=1,
            tags=("round",),
        )

        # 让内部内容跟随尺寸变化
        inner_w = max(w - self.pad_x * 2, 10)
        inner_h = max(h - self.pad_y * 2, 10)
        self.coords(self._win, self.pad_x, self.pad_y)
        self.itemconfigure(self._win, width=inner_w, height=inner_h)


class IconButton(tk.Canvas):
    """自绘小图标按钮（眼睛/复制），用于输入框内部。"""

    def __init__(self, master, *, kind: str, command, size: int = 18, bg: str = "#ffffff"):
        super().__init__(master, width=size, height=size, highlightthickness=0, bd=0, bg=bg)
        self.kind = kind
        self.command = command
        self.size = size
        self._eye_open = True
        self._hover = False
        self._draw()

        self.bind("<Button-1>", lambda _e: self.command())
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_enter(self, _e):
        self._hover = True
        self.configure(cursor="hand2")
        self._draw()

    def _on_leave(self, _e):
        self._hover = False
        self.configure(cursor="")
        self._draw()

    def set_eye_open(self, eye_open: bool) -> None:
        self._eye_open = bool(eye_open)
        if self.kind == "eye":
            self._draw()

    def _color(self) -> str:
        return "#111827" if self._hover else "#374151"

    def _draw(self) -> None:
        self.delete("all")
        c = self._color()
        s = self.size

        if self.kind == "copy":
            # 两个重叠的矩形代表复制
            self.create_rectangle(6, 5, s - 4, s - 6, outline=c, width=1)
            self.create_rectangle(4, 7, s - 6, s - 4, outline=c, width=1)
            return

        if self.kind == "eye":
            # 简单眼睛：外轮廓 + 瞳孔
            self.create_oval(3, 6, s - 3, s - 6, outline=c, width=1)
            self.create_oval(s / 2 - 2, s / 2 - 2, s / 2 + 2, s / 2 + 2, fill=c, outline=c)
            if not self._eye_open:
                # 关闭：画一条斜杠
                self.create_line(4, s - 5, s - 4, 5, fill=c, width=2)
            return

try:
    import socks  # PySocks，测试代理时经 SOCKS5 访问 ipinfo；显式 import 便于 PyInstaller 打包
    _HAS_PYSOCKS = True
except ImportError:
    _HAS_PYSOCKS = False

try:
    import openpyxl  # 导入 Excel (.xlsx) 时使用
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

PROXY_TEST_URL = "https://ipinfo.io/json"

# 已保存数据的存储路径（与脚本或 exe 同目录）
def _app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# 新文件名
SAVED_EMAIL_PATH = os.path.join(_app_dir(), "saved_emails.json")
SAVED_PROXIES_PATH = os.path.join(_app_dir(), "saved_proxies.json")
SAVED_ACCOUNTS_PATH = os.path.join(_app_dir(), "saved_accounts.json")
SAVED_BROWSER_PROFILES_PATH = os.path.join(_app_dir(), "saved_browser_profiles.json")
# 首页“一键启动”综合数据（账号+邮箱+代理+浏览器）
SAVED_LAUNCH_PROFILES_PATH = os.path.join(_app_dir(), "saved_launch_profiles.json")


def _load_json_list(path: str) -> list[dict]:
    """读取 list[dict]（不做兼容迁移）。"""
    try:
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


def _save_json_list(path: str, items: list[dict]) -> None:
    """写入 list[dict]（不做兼容迁移）。"""
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_saved_accounts() -> list[dict]:
    """加载已保存的账号列表。每项: email, password, imap(可选)"""
    return _load_json_list(SAVED_EMAIL_PATH)


def save_saved_accounts(accounts: list[dict]) -> None:
    """将账号列表写入本地 JSON"""
    _save_json_list(SAVED_EMAIL_PATH, accounts)


def load_saved_proxies() -> list[dict]:
    """加载已保存的代理配置列表。每项: forward, share_url, proxy_port（按share_url标识）"""
    return _load_json_list(SAVED_PROXIES_PATH)


def save_saved_proxies(proxies: list[dict]) -> None:
    """将代理配置列表写入本地 JSON"""
    _save_json_list(SAVED_PROXIES_PATH, proxies)


def load_saved_fb_accounts() -> list[dict]:
    """
    加载已保存的“扩展账号信息”(例如 FB 注册信息)。
    每项示例：
      {
        "cookies": [],
        "email": "...",
        "email_password": "...",
        "password": "...",
        "account_id": "...",
        "regDate": "YYYY-MM-DD",
        "share_url": "socks5://..."
      }
    """
    return _load_json_list(SAVED_ACCOUNTS_PATH)


def save_saved_fb_accounts(accounts: list[dict]) -> None:
    """将扩展账号列表写入本地 JSON"""
    _save_json_list(SAVED_ACCOUNTS_PATH, accounts)


def load_saved_browser_profiles() -> list[dict]:
    """加载已保存的浏览器启动配置（指纹参数等）。"""
    try:
        if os.path.isfile(SAVED_BROWSER_PROFILES_PATH):
            with open(SAVED_BROWSER_PROFILES_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
    except Exception:
        pass
    return []


def save_saved_browser_profiles(profiles: list[dict]) -> None:
    """保存浏览器启动配置列表到本地 JSON。"""
    try:
        with open(SAVED_BROWSER_PROFILES_PATH, "w", encoding="utf-8") as f:
            json.dump(profiles, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def load_saved_launch_profiles() -> list[dict]:
    """加载首页“一键启动”的综合配置列表。"""
    return _load_json_list(SAVED_LAUNCH_PROFILES_PATH)


def save_saved_launch_profiles(items: list[dict]) -> None:
    """保存首页“一键启动”的综合配置列表。"""
    _save_json_list(SAVED_LAUNCH_PROFILES_PATH, items)


def _launch_key_of_account(a: dict) -> str:
    """为账号生成稳定 key（用于综合配置的去重更新）。"""
    plat = str(a.get("platform") or "").strip().lower()
    aid = str(a.get("account_id") or "").strip()
    email = str(a.get("email") or "").strip().lower()
    # platform + (account_id 优先，否则 email)
    core = aid or email
    return f"{plat}::{core}" if core else f"{plat}::(empty)"


def _build_launch_profile_for_account(
    *,
    account: dict,
    emails: list[dict] | None = None,
    proxies: list[dict] | None = None,
    browser_profiles: list[dict] | None = None,
) -> dict:
    """把账号信息拼成“综合数据”，用于首页一键启动。"""
    a = account if isinstance(account, dict) else {}
    email_addr = str(a.get("email") or "").strip()
    share_url = str(a.get("share_url") or "").strip()
    browser_name = str(a.get("browser_profile") or "").strip()

    # 1) 邮箱信息：优先取 saved_emails.json 里同邮箱的记录；否则用账号里的 email_password
    emails = emails if isinstance(emails, list) else []
    email_rec: dict = {}
    em_lower = email_addr.lower()
    for it in emails:
        if not isinstance(it, dict):
            continue
        if str(it.get("email") or "").strip().lower() == em_lower and em_lower:
            email_rec = dict(it)
            break
    if not email_rec:
        email_rec = {
            "email": email_addr,
            "password": str(a.get("email_password") or "").strip(),
            "imap": "",
        }

    # 2) 代理信息：按 share_url 在 saved_proxies.json 中找；找不到就只保留 share_url（让首页仍可回填）
    proxies = proxies if isinstance(proxies, list) else []
    proxy_rec: dict = {"share_url": share_url}
    su = share_url.strip()
    for it in proxies:
        if not isinstance(it, dict):
            continue
        if str(it.get("share_url") or "").strip() == su and su:
            proxy_rec = dict(it)
            break

    # 3) 浏览器信息：按 browser_profile 名称在 saved_browser_profiles.json 中找
    browser_profiles = browser_profiles if isinstance(browser_profiles, list) else []
    browser_rec: dict = {"name": browser_name} if browser_name else {}
    for it in browser_profiles:
        if not isinstance(it, dict):
            continue
        if str(it.get("name") or "").strip() == browser_name and browser_name:
            browser_rec = dict(it)
            break

    return {
        "key": _launch_key_of_account(a),
        "platform": str(a.get("platform") or "").strip(),
        "account_id": str(a.get("account_id") or "").strip(),
        "email": email_addr,
        "share_url": share_url,
        "browser_profile": browser_name,
        "account": dict(a),
        "email_info": email_rec,
        "proxy_info": proxy_rec,
        "browser_info": browser_rec,
        "updatedAt": _dt.datetime.now().isoformat(timespec="seconds"),
    }


# 常见邮箱 IMAP 服务器（域名后缀 -> (host, port)）
IMAP_PRESETS = {
    "gmail.com": ("imap.gmail.com", 993),
    "googlemail.com": ("imap.gmail.com", 993),
    "outlook.com": ("outlook.office365.com", 993),
    "hotmail.com": ("outlook.office365.com", 993),
    "live.com": ("outlook.office365.com", 993),
    "msn.com": ("outlook.office365.com", 993),
    "office365.com": ("outlook.office365.com", 993),
    "qq.com": ("imap.qq.com", 993),
    "foxmail.com": ("imap.qq.com", 993),
    "163.com": ("imap.163.com", 993),
    "126.com": ("imap.126.com", 993),
    "yeah.net": ("imap.yeah.net", 993),
    "sina.com": ("imap.sina.com", 993),
    "sina.cn": ("imap.sina.cn", 993),
    "sohu.com": ("imap.sohu.com", 993),
    "yahoo.com": ("imap.mail.yahoo.com", 993),
    "yahoo.cn": ("imap.mail.yahoo.com", 993),
    "icloud.com": ("imap.mail.me.com", 993),
    "me.com": ("imap.mail.me.com", 993),
    "mail.com": ("imap.mail.com", 993),
    "aol.com": ("imap.aol.com", 993),
    "protonmail.com": ("127.0.0.1", 1143),  # Proton 需桥接，此处示意
    "yandex.com": ("imap.yandex.com", 993),
    "zoho.com": ("imap.zoho.com", 993),
}

# 常见 SMTP 服务器（域名后缀 -> (host, port)）；未列出的按 smtp.{域名}:587 生成
SMTP_PRESETS = {
    "gmail.com": ("smtp.gmail.com", 587),
    "googlemail.com": ("smtp.gmail.com", 587),
    "outlook.com": ("smtp.office365.com", 587),
    "hotmail.com": ("smtp.office365.com", 587),
    "live.com": ("smtp.office365.com", 587),
    "office365.com": ("smtp.office365.com", 587),
    "qq.com": ("smtp.qq.com", 587),
    "foxmail.com": ("smtp.qq.com", 587),
    "163.com": ("smtp.163.com", 465),
    "126.com": ("smtp.126.com", 465),
    "yahoo.com": ("smtp.mail.yahoo.com", 587),
    "icloud.com": ("smtp.mail.me.com", 587),
}


def _find_col(hdr_lower: list[str], names: list[str]) -> int | None:
    """在表头中按名称找列索引，精确匹配。names 为候选列名（小写）。"""
    for i, h in enumerate(hdr_lower):
        if h in names:
            return i
    return None


def _domain(addr: str) -> str | None:
    """从邮箱地址提取域名，无效则返回 None"""
    addr = (addr or "").strip().lower()
    if "@" in addr:
        return addr.split("@", 1)[1]
    return None


def guess_imap_server(addr: str) -> tuple[str, int]:
    """
    根据邮箱地址推断 IMAP 服务器。
    先在预设表中查找；若未匹配，则按域名自动生成 imap.{域名}，如 imap.retrorabbit.net。
    """
    domain = _domain(addr)
    if domain:
        for suffix, (host, port) in IMAP_PRESETS.items():
            if domain == suffix or domain.endswith("." + suffix):
                return host, port
        return f"imap.{domain}", 993
    return "imap.gmail.com", 993  # 无法解析域名时的默认


def guess_smtp_server(addr: str) -> tuple[str, int]:
    """
    根据邮箱地址推断 SMTP 服务器。
    先在预设表中查找；若未匹配，则按域名自动生成 smtp.{域名}，如 smtp.retrorabbit.net。
    端口默认 587（Submission）。
    """
    domain = _domain(addr)
    if domain:
        for suffix, (host, port) in SMTP_PRESETS.items():
            if domain == suffix or domain.endswith("." + suffix):
                return host, port
        return f"smtp.{domain}", 587
    return "smtp.gmail.com", 587  # 无法解析时的默认


def decode_mime_header(s: str | None) -> str:
    """解析 MIME 编码的邮件头"""
    if not s:
        return ""
    try:
        parts = decode_header(s)
        buf = []
        for part, enc in parts:
            if isinstance(part, bytes):
                buf.append(part.decode(enc or "utf-8", errors="replace"))
            else:
                buf.append(str(part))
        return "".join(buf).strip()
    except Exception:
        return str(s) if s else ""


def html_to_plain(html: str) -> str:
    """将 HTML 转为纯文本，避免展示源码"""
    if not (html or "").strip():
        return ""
    s = re.sub(r"<script[^>]*>[\s\S]*?</script>", " ", html, flags=re.I)
    s = re.sub(r"<style[^>]*>[\s\S]*?</style>", " ", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = re.sub(r"&nbsp;", " ", s, flags=re.I)
    s = re.sub(r"&lt;", "<", s, flags=re.I)
    s = re.sub(r"&gt;", ">", s, flags=re.I)
    s = re.sub(r"&amp;", "&", s, flags=re.I)
    s = re.sub(r"&quot;", '"', s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def get_email_bodies(msg: email.message.Message, max_plain: int = 1200, max_html: int = 2400) -> tuple[str, str]:
    """
    提取邮件正文：纯文本 与 HTML。返回 (body_plain, body_html)。
    若仅有 HTML，则 body_plain 用 html_to_plain(html) 得到。
    """
    body_plain = ""
    body_html = ""

    def _get(m: email.message.Message) -> None:
        nonlocal body_plain, body_html
        if m.get("Content-Disposition") is not None:
            return
        ctype = (m.get_content_type() or "").lower()
        enc = m.get_content_charset() or "utf-8"
        try:
            raw = m.get_payload(decode=True)
            text = (raw.decode(enc, errors="replace") if raw else "") or ""
        except Exception:
            return
        if "text/plain" in ctype:
            body_plain = re.sub(r"\s+", " ", (text or "").strip())
            if len(body_plain) > max_plain:
                body_plain = body_plain[:max_plain] + "..."
        elif "text/html" in ctype:
            body_html = (text or "").strip()
            if len(body_html) > max_html:
                body_html = body_html[:max_html] + "..."

    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_maintype() == "text":
                _get(part)
    else:
        if msg.get_content_maintype() == "text":
            _get(msg)

    if not body_plain and body_html:
        body_plain = html_to_plain(body_html)
        if len(body_plain) > max_plain:
            body_plain = body_plain[:max_plain] + "..."
    return body_plain, body_html


def extract_verification_code(text: str) -> str:
    """
    从邮件正文中提取验证码。常见：4～8 位数字、验证码/code 后的数字等。
    """
    if not (text or "").strip():
        return ""
    # 0. Facebook 常见格式：FB-12345 / FB‑12345 / FB-12345Don't...
    #    - 先匹配，避免被正文里的其它数字（如邮编）误命中
    #    - 返回数字部分，便于直接粘贴到验证码框
    #    - 兼容不同短横线：- ‑ – —
    fb_pat = re.compile(r"(?i)(?<!\d)fb[\-\u2011\u2013\u2014_\s]*([0-9]{4,8})(?!\d)")
    fb_hits = [(m.start(), m.group(1)) for m in fb_pat.finditer(text)]
    if fb_hits:
        if len(fb_hits) == 1:
            return fb_hits[0][1].strip()
        # 多次出现时，优先选择靠近 “confirmation code / code / 验证码” 的那一个
        best_code = fb_hits[0][1]
        best_score = -1
        tl = text.lower()
        for pos, code in fb_hits:
            window = tl[pos : pos + 120]
            score = 0
            if "confirmation code" in window:
                score += 3
            if "verification code" in window:
                score += 2
            if " code" in window:
                score += 1
            if "验证码" in window:
                score += 2
            if score > best_score:
                best_score = score
                best_code = code
        return (best_code or "").strip()
    # 1. 常见表述：验证码/code/verification code 后的数字（含 "enter this code in X: 123456"）
    for pat in [
        r"(?:验证码|校验码|动态码|授权码|security\s*code|verification\s*code|code|你的码)[：:\s]*[是为]?\s*(\d{4,8})",
        r"(?:verification\s*code|code)[^0-9]{0,100}(\d{4,8})\b",
        r"(\d{4,8})\s*(?:是您的验证码|为您的验证码|为验证码)",
        r"[\（\(](\d{4,8})[\）\)]",
    ]:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1).strip()
    # 2. 纯数字 4～8 位（优先 5～6 位）
    for w in [6, 5, 8, 7, 4]:
        m = re.search(rf"\b(\d{{{w}}})\b", text)
        if m:
            return m.group(1)
    m = re.search(r"\b(\d{4,8})\b", text)
    return m.group(1) if m else ""


def fetch_recent_emails(email_addr: str, password: str, limit: int = 3, imap_host: str | None = None, imap_port: int = 993) -> list[dict] | str:
    """
    通过 IMAP 登录并获取最近 limit 封邮件。
    成功返回 list[dict]，每个 dict: subject, from_addr, date, body_preview, raw_date
    失败返回错误信息 str。
    """
    host = imap_host or guess_imap_server(email_addr)[0]
    if imap_port <= 0:
        imap_port = 993

    try:
        with imaplib.IMAP4_SSL(host, port=imap_port, timeout=30) as imap:
            imap.login(email_addr, password)
            imap.select("INBOX", readonly=True)
            # 按日期倒序，搜索全部
            status, data = imap.search(None, "ALL")
            if status != "OK" or not data:
                return []

            ids = data[0].split()
            ids = ids[-limit:] if len(ids) >= limit else ids
            ids.reverse()  # 最新的在前

            out = []
            for bid in ids:
                try:
                    status, raw = imap.fetch(bid, "(RFC822)")
                    if status != "OK" or not raw:
                        continue
                    msg = email.message_from_bytes(raw[0][1])
                    subject = decode_mime_header(msg.get("Subject"))
                    from_h = decode_mime_header(msg.get("From"))
                    date_h = msg.get("Date") or ""
                    body_plain, body_html = get_email_bodies(msg)
                    # 从纯文本提取验证码（body_plain 已含 html_to_plain 的结果）
                    code = extract_verification_code(body_plain)
                    # 展示用：纯文本优先，无则从 HTML 转
                    body_display = body_plain or html_to_plain(body_html or "")
                    out.append({
                        "subject": subject,
                        "from_addr": from_h,
                        "date": date_h,
                        "body_display": body_display,
                        "code": code,
                    })
                except Exception:
                    continue

            return out
    except imaplib.IMAP4.error as e:
        return f"IMAP 登录或读取失败: {e}"
    except OSError as e:
        return f"网络或连接错误: {e}"
    except Exception as e:
        return f"错误: {e}"


class EmailInboxReaderApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("收件箱 - 最近 5 封邮件")
        self.root.geometry("760x780")
        self.root.resizable(True, True)

        self._accounts = load_saved_accounts()
        self._proxies = load_saved_proxies()  # 已保存的代理配置
        self._gost_proc = None  # 代理子进程
        self._pwd_visible = False
        self._build_ui()

    def _close_browser(self):
        """一键关闭：尽量关闭最近一次启动的浏览器进程树。"""
        proc = getattr(self, "_browser_proc", None)
        pid = None
        try:
            if proc is not None and getattr(proc, "pid", None):
                pid = int(proc.pid)
        except Exception:
            pid = None
        if not pid:
            pid = getattr(self, "_browser_last_pid", None)
        try:
            pid = int(pid) if pid else None
        except Exception:
            pid = None
        if not pid:
            return
        try:
            subprocess.Popen(
                ["taskkill", "/PID", str(pid), "/T", "/F"],
                creationflags=0x08000000 if sys.platform == "win32" else 0,  # CREATE_NO_WINDOW
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception:
            try:
                if proc is not None:
                    proc.terminate()
            except Exception:
                pass

    def _sync_proxy_port_into_others(self, other_text: str, port: str) -> str:
        """把 others 里的 --proxy-server=...127.0.0.1:PORT 同步到指定 port。"""
        p = (port or "").strip()
        if not p.isdigit():
            return other_text
        s = other_text or ""
        # 1) 带引号
        s = re.sub(
            r'(--proxy-server=)(["\'])(socks5://127\.0\.0\.1:)(\d+)(\2)',
            rf"\1\2\3{p}\5",
            s,
            flags=re.I,
        )
        # 2) 不带引号
        s = re.sub(
            r'(--proxy-server=)(socks5://127\.0\.0\.1:)(\d+)',
            rf"\1\2{p}",
            s,
            flags=re.I,
        )
        return s

    def _one_click_start(self, account: dict):
        """首页“一键启动”：加载邮箱/代理/浏览器并自动开启代理+启动浏览器。"""
        a = account if isinstance(account, dict) else {}
        emails_all = load_saved_accounts()
        proxies_all = load_saved_proxies()
        browsers_all = load_saved_browser_profiles()
        lp = _build_launch_profile_for_account(account=a, emails=emails_all, proxies=proxies_all, browser_profiles=browsers_all)

        email_info = lp.get("email_info") if isinstance(lp.get("email_info"), dict) else {}
        proxy_info = lp.get("proxy_info") if isinstance(lp.get("proxy_info"), dict) else {}
        browser_info = lp.get("browser_info") if isinstance(lp.get("browser_info"), dict) else {}

        # 1) 收件箱：邮箱
        try:
            self.var_email.set(str(email_info.get("email") or a.get("email") or "").strip())
            self.var_pwd.set(str(email_info.get("password") or a.get("email_password") or "").strip())
            self.var_imap.set(str(email_info.get("imap") or "").strip())
        except Exception:
            pass

        # 2) 收件箱：代理（刷新下拉并联动 forward/port）
        try:
            self._proxies = load_saved_proxies()
            self._refresh_proxy_combo()
        except Exception:
            pass
        share_url = str(proxy_info.get("share_url") or a.get("share_url") or "").strip()
        if share_url:
            try:
                self.var_share_url.set(share_url)
            except Exception:
                pass
        try:
            fwd = str(proxy_info.get("forward") or "").strip()
            if fwd:
                self.var_forward.set(fwd)
            pp = str(proxy_info.get("proxy_port") or "").strip()
            if pp:
                self.var_proxy_port.set(pp)
        except Exception:
            pass
        try:
            self._on_proxy_select()
        except Exception:
            pass

        # 3) 浏览器：应用环境
        try:
            bp_name = (str(a.get("browser_profile") or "").strip() or str(browser_info.get("name") or "").strip())
            if bp_name and hasattr(self, "_browser_apply_profile_by_name"):
                self._browser_apply_profile_by_name(bp_name)
            elif browser_info and hasattr(self, "_browser_apply_profile"):
                self._browser_apply_profile(browser_info)
        except Exception:
            pass

        # 4) 同步代理端口到 others
        try:
            if hasattr(self, "_browser_txt_other") and self._browser_txt_other is not None:
                cur = self._browser_txt_other.get("1.0", tk.END).strip()
                new = self._sync_proxy_port_into_others(cur, (self.var_proxy_port.get() or "").strip())
                if new != cur:
                    self._browser_txt_other.delete("1.0", tk.END)
                    self._browser_txt_other.insert("1.0", new)
            if hasattr(self, "_browser_update_preview"):
                self._browser_update_preview()
        except Exception:
            pass

        # 5) 自动开启代理（先断开再开启，避免“已在运行”）
        try:
            self._on_stop_proxy()
        except Exception:
            pass
        try:
            self._on_start_proxy()
        except Exception:
            pass

        # 6) 启动浏览器
        try:
            if hasattr(self, "_browser_launch"):
                self._browser_launch()
        except Exception:
            pass

    def _one_click_load(self, account: dict):
        """首页“一键加载”：仅加载邮箱/代理/浏览器配置，不启动代理和浏览器。"""
        a = account if isinstance(account, dict) else {}
        emails_all = load_saved_accounts()
        proxies_all = load_saved_proxies()
        browsers_all = load_saved_browser_profiles()
        lp = _build_launch_profile_for_account(account=a, emails=emails_all, proxies=proxies_all, browser_profiles=browsers_all)

        email_info = lp.get("email_info") if isinstance(lp.get("email_info"), dict) else {}
        proxy_info = lp.get("proxy_info") if isinstance(lp.get("proxy_info"), dict) else {}
        browser_info = lp.get("browser_info") if isinstance(lp.get("browser_info"), dict) else {}

        # 1) 收件箱：邮箱
        try:
            self.var_email.set(str(email_info.get("email") or a.get("email") or "").strip())
            self.var_pwd.set(str(email_info.get("password") or a.get("email_password") or "").strip())
            self.var_imap.set(str(email_info.get("imap") or "").strip())
        except Exception:
            pass

        # 2) 收件箱：代理（刷新下拉并联动 forward/port）
        try:
            self._proxies = load_saved_proxies()
            self._refresh_proxy_combo()
        except Exception:
            pass
        share_url = str(proxy_info.get("share_url") or a.get("share_url") or "").strip()
        if share_url:
            try:
                self.var_share_url.set(share_url)
            except Exception:
                pass
        try:
            fwd = str(proxy_info.get("forward") or "").strip()
            if fwd:
                self.var_forward.set(fwd)
            pp = str(proxy_info.get("proxy_port") or "").strip()
            if pp:
                self.var_proxy_port.set(pp)
        except Exception:
            pass
        try:
            self._on_proxy_select()
        except Exception:
            pass

        # 3) 浏览器：应用环境
        try:
            bp_name = (str(a.get("browser_profile") or "").strip() or str(browser_info.get("name") or "").strip())
            if bp_name and hasattr(self, "_browser_apply_profile_by_name"):
                self._browser_apply_profile_by_name(bp_name)
            elif browser_info and hasattr(self, "_browser_apply_profile"):
                self._browser_apply_profile(browser_info)
        except Exception:
            pass

        # 4) 同步代理端口到 others + 刷新预览
        try:
            if hasattr(self, "_browser_txt_other") and self._browser_txt_other is not None:
                cur = self._browser_txt_other.get("1.0", tk.END).strip()
                new = self._sync_proxy_port_into_others(cur, (self.var_proxy_port.get() or "").strip())
                if new != cur:
                    self._browser_txt_other.delete("1.0", tk.END)
                    self._browser_txt_other.insert("1.0", new)
            if hasattr(self, "_browser_update_preview"):
                self._browser_update_preview()
        except Exception:
            pass

    def _build_home_tab(self, parent: tk.Widget):
        """首页：统计 + 选择账号 + 一键启动。"""
        wrap = ttk.Frame(parent, padding=12)
        wrap.pack(fill=tk.BOTH, expand=True)

        # 顶部酷炫头图（渐变 + 轻动画）
        header = tk.Canvas(wrap, height=120, highlightthickness=0)
        header.pack(fill=tk.X, pady=(0, 10))
        header.configure(bg="#0b1220")

        def _draw_gradient():
            header.delete("all")
            w = max(1, header.winfo_width())
            h = max(1, header.winfo_height())
            # 简单横向渐变：深蓝 -> 蓝绿
            left = (11, 18, 32)
            right = (14, 165, 233)
            steps = 60
            for i in range(steps):
                t = i / max(1, steps - 1)
                r = int(left[0] + (right[0] - left[0]) * t)
                g = int(left[1] + (right[1] - left[1]) * t)
                b = int(left[2] + (right[2] - left[2]) * t)
                x0 = int(w * i / steps)
                x1 = int(w * (i + 1) / steps) + 1
                header.create_rectangle(x0, 0, x1, h, outline="", fill=f"#{r:02x}{g:02x}{b:02x}")
            header.create_text(18, 34, anchor="w", text="首页 / 控制台", fill="white", font=("Segoe UI", 18, "bold"))
            header.create_text(18, 72, anchor="w", text="平台账号统计 · 一键启动（自动加载邮箱/代理/浏览器）", fill="#dbeafe", font=("Segoe UI", 10))

        dots = [{"x": 40, "y": 25, "vx": 1.2}, {"x": 120, "y": 90, "vx": 0.9}, {"x": 240, "y": 55, "vx": 1.5}]

        def _animate():
            try:
                _draw_gradient()
                w = max(1, header.winfo_width())
                for d in dots:
                    d["x"] += d["vx"]
                    if d["x"] > w + 40:
                        d["x"] = -40
                    header.create_oval(d["x"] - 6, d["y"] - 6, d["x"] + 6, d["y"] + 6, outline="", fill="#93c5fd")
                    header.create_oval(d["x"] - 12, d["y"] - 12, d["x"] + 12, d["y"] + 12, outline="", fill="#93c5fd", stipple="gray25")
            except Exception:
                pass
            self.root.after(80, _animate)

        header.bind("<Configure>", lambda _e: _draw_gradient())
        _animate()

        # 统计卡片区
        stats_bar = ttk.Frame(wrap)
        stats_bar.pack(fill=tk.X, pady=(0, 10))

        # 账号列表 + 详情
        body = ttk.Frame(wrap)
        body.pack(fill=tk.BOTH, expand=True)
        # 左右各占一半
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        left = ttk.Frame(body)
        left.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 6))
        right = ttk.Frame(body)
        right.grid(row=0, column=1, sticky=tk.NSEW, padx=(6, 0))

        # 筛选
        filt = ttk.Frame(left)
        filt.pack(fill=tk.X, pady=(0, 6))
        var_plat = tk.StringVar(value="全部")
        ttk.Label(filt, text="平台").pack(side=tk.LEFT)
        combo_plat = ttk.Combobox(filt, textvariable=var_plat, width=12, state="readonly")
        combo_plat.pack(side=tk.LEFT, padx=(6, 10))
        ttk.Label(filt, text="搜索").pack(side=tk.LEFT)
        var_kw = tk.StringVar(value="")
        ent_kw = ttk.Entry(filt, textvariable=var_kw, width=26)
        ent_kw.pack(side=tk.LEFT, padx=(6, 10))

        btn_refresh = ttk.Button(filt, text="刷新")
        btn_refresh.pack(side=tk.RIGHT)

        # 列表（左侧只展示平台 + 邮箱）
        cols = ("platform", "email")
        tree = ttk.Treeview(left, columns=cols, show="headings", selectmode="browse", height=17)
        vsb = ttk.Scrollbar(left, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        tree.heading("platform", text="平台")
        tree.heading("email", text="邮箱")
        tree.column("platform", width=110, anchor="w")
        tree.column("email", width=260, anchor="w")

        # 详情 + 一键启动
        detail = ttk.LabelFrame(right, text="详情", padding=10)
        detail.pack(fill=tk.BOTH, expand=True)
        txt_detail = scrolledtext.ScrolledText(detail, height=12, font=("Consolas", 10))
        txt_detail.pack(fill=tk.BOTH, expand=True)
        txt_detail.configure(state="disabled")

        btns = ttk.Frame(right)
        btns.pack(fill=tk.X, pady=(10, 0))
        btn_start = ttk.Button(btns, text="一键启动")
        btn_start.pack(side=tk.LEFT)
        btn_load = ttk.Button(btns, text="一键加载")
        btn_load.pack(side=tk.LEFT, padx=(10, 0))
        btn_close = ttk.Button(btns, text="一键关闭")
        btn_close.pack(side=tk.LEFT, padx=(10, 0))

        iid_to_account: dict[str, dict] = {}

        def _platform_display(s: str) -> str:
            s = (s or "").strip()
            return s if s else "未设置"

        def _render_detail(a: dict):
            txt_detail.configure(state="normal")
            txt_detail.delete("1.0", tk.END)
            if not a:
                txt_detail.insert("1.0", "（未选择账号）")
            else:
                lines = [
                    f"平台: {_platform_display(str(a.get('platform') or ''))}",
                    f"邮箱: {str(a.get('email') or '').strip()}",
                    f"邮箱密码: {str(a.get('email_password') or '').strip()}",
                    f"账号ID: {str(a.get('account_id') or '').strip()}",
                    f"代理: {str(a.get('share_url') or '').strip()}",
                    f"浏览器环境: {str(a.get('browser_profile') or '').strip()}",
                    f"注册日期: {str(a.get('regDate') or '').strip()}",
                ]
                txt_detail.insert("1.0", "\n".join(lines))
            txt_detail.configure(state="disabled")

        def _render_stats(accounts: list[dict]):
            for w in stats_bar.winfo_children():
                w.destroy()
            counts: dict[str, int] = {}
            for a in accounts:
                if not isinstance(a, dict):
                    continue
                p = _platform_display(str(a.get("platform") or ""))
                counts[p] = counts.get(p, 0) + 1
            if not counts:
                ttk.Label(stats_bar, text="暂无账号数据（去“账号管理”保存一些账号）").pack(side=tk.LEFT)
                return
            palette = ["#0ea5e9", "#22c55e", "#a855f7", "#f97316", "#ef4444", "#14b8a6"]
            for i, (p, n) in enumerate(sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[:6]):
                bg = palette[i % len(palette)]
                card = tk.Frame(stats_bar, bg=bg, padx=12, pady=10)
                card.pack(side=tk.LEFT, padx=(0, 10))
                tk.Label(card, text=p, bg=bg, fg="white", font=("Segoe UI", 10, "bold")).pack(anchor="w")
                tk.Label(card, text=str(n), bg=bg, fg="white", font=("Segoe UI", 18, "bold")).pack(anchor="w")

        def _refresh():
            raw = load_saved_fb_accounts()
            arr = [a for a in raw if isinstance(a, dict)]

            # 平台下拉
            plats = sorted({_platform_display(str(a.get("platform") or "")) for a in arr})
            combo_plat["values"] = ["全部"] + plats
            if var_plat.get() not in combo_plat["values"]:
                var_plat.set("全部")

            # 统计
            _render_stats(arr)

            # 列表
            tree.delete(*tree.get_children())
            iid_to_account.clear()
            fp = var_plat.get()
            kw = (var_kw.get() or "").strip().lower()
            for a in arr:
                p = _platform_display(str(a.get("platform") or ""))
                if fp and fp != "全部" and p != fp:
                    continue
                em = str(a.get("email") or "").strip()
                aid = str(a.get("account_id") or "").strip()
                su = str(a.get("share_url") or "").strip()
                bp = str(a.get("browser_profile") or "").strip()
                blob = " ".join([p, em, aid, su, bp]).lower()
                if kw and kw not in blob:
                    continue
                iid = tree.insert("", "end", values=(p, em))
                iid_to_account[str(iid)] = a
            _render_detail({})

        def _on_select(_event=None):
            sel = tree.selection()
            if not sel:
                _render_detail({})
                return
            a = iid_to_account.get(str(sel[0]), {})
            _render_detail(a)

        def _start():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("未选择", "请先选择一个账号再一键启动。", parent=self.root)
                return
            a = iid_to_account.get(str(sel[0]), {})
            if not isinstance(a, dict) or not (a.get("email") or "").strip():
                messagebox.showwarning("数据不完整", "该账号缺少邮箱，无法一键启动。", parent=self.root)
                return
            self._one_click_start(a)

        def _load():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("未选择", "请先选择一个账号再一键加载。", parent=self.root)
                return
            a = iid_to_account.get(str(sel[0]), {})
            if not isinstance(a, dict) or not (a.get("email") or "").strip():
                messagebox.showwarning("数据不完整", "该账号缺少邮箱，无法一键加载。", parent=self.root)
                return
            self._one_click_load(a)

        def _close():
            try:
                self._on_stop_proxy()
            except Exception:
                pass
            try:
                self._close_browser()
            except Exception:
                pass

        btn_refresh.configure(command=_refresh)
        combo_plat.bind("<<ComboboxSelected>>", lambda _e: _refresh())
        ent_kw.bind("<Return>", lambda _e: _refresh())
        tree.bind("<<TreeviewSelect>>", _on_select)
        btn_start.configure(command=_start)
        btn_load.configure(command=_load)
        btn_close.configure(command=_close)

        # 暴露刷新函数：账号管理删除/保存后可自动刷新首页
        self._home_refresh = _refresh

        _refresh()
        self._refresh_account_combo()
        self._refresh_proxy_combo()

    def _log(self, msg: str):
        def _do():
            self.log.insert(tk.END, msg + "\n")
            self.log.see(tk.END)
        self.root.after(0, _do)

    def _set_loading(self, loading: bool):
        def _do():
            self.btn_login.config(state=tk.DISABLED if loading else tk.NORMAL)
            self.lbl_status.config(text="正在登录并读取…" if loading else "就绪")
        self.root.after(0, _do)

    def _refresh_account_combo(self):
        emails = sorted({(a.get("email") or "").strip() for a in self._accounts if (a.get("email") or "").strip()})
        self._all_account_emails = emails
        self.combo_email["values"] = emails
        if not emails:
            self.combo_email.set("")

    def _find_account_by_email(self, email_text: str) -> dict | None:
        t = (email_text or "").strip().lower()
        if not t:
            return None
        for a in self._accounts:
            em = (a.get("email") or "").strip().lower()
            if em == t:
                return a
        return None

    def _try_fill_account_fields(self, email_text: str) -> None:
        """当邮箱命中已保存账号时，自动回填密码与 IMAP。"""
        a = self._find_account_by_email(email_text)
        if not a:
            return
        self.var_pwd.set(a.get("password", ""))
        self.var_imap.set(a.get("imap", ""))

    def _filter_account_combo(self, query: str) -> None:
        """输入即过滤邮箱下拉列表（不改变当前输入内容）。"""
        q = (query or "").strip().lower()
        all_emails = getattr(self, "_all_account_emails", None)
        if all_emails is None:
            all_emails = sorted({(a.get("email") or "").strip() for a in self._accounts if (a.get("email") or "").strip()})
            self._all_account_emails = all_emails
        if not q:
            values = all_emails
        else:
            values = [e for e in all_emails if q in e.lower()]
        self.combo_email["values"] = values
        # 尽量弹出下拉候选（某些主题/系统可能不支持，失败则忽略）
        if values:
            try:
                self.combo_email.tk.call("ttk::combobox::Post", self.combo_email)
            except Exception:
                pass

    def _on_email_keyrelease(self, event=None):
        # 过滤用：忽略方向键/回车等控制按键
        if event is not None and getattr(event, "keysym", "") in {"Up", "Down", "Left", "Right", "Return", "Escape", "Tab"}:
            return
        self._filter_account_combo(self.var_email.get())

    def _on_email_commit(self, event=None):
        """回车/失焦时：若输入命中已保存账号则自动回填。"""
        self._try_fill_account_fields(self.var_email.get())

    def _on_email_var_changed(self, *_):
        """任意输入变化：若已精确匹配到保存邮箱，立即回填。"""
        self._try_fill_account_fields(self.var_email.get())

    def _refresh_proxy_combo(self):
        """刷新代理配置下拉框的选项（按share_url）"""
        share_urls = [p.get("share_url", "") for p in self._proxies if p.get("share_url")]
        self.combo_proxy["values"] = share_urls
        if not share_urls:
            self.combo_proxy.set("")

    def _on_proxy_select(self, event=None):
        """代理配置下拉选已保存配置时，自动填充转发地址、本地端口"""
        sel = (self.var_share_url.get() or "").strip()
        if not sel:
            return
        for p in self._proxies:
            if (p.get("share_url") or "").strip() == sel:
                self.var_forward.set(p.get("forward", "socks5://127.0.0.1:10809"))
                self.var_proxy_port.set(p.get("proxy_port", "50681"))
                return

    def _save_current_proxy(self) -> bool:
        """保存当前表单中的代理配置到列表并持久化。已存在则更新。返回是否成功。"""
        forward = (self.var_forward.get() or "").strip()
        share_url = (self.var_share_url.get() or "").strip()
        proxy_port = (self.var_proxy_port.get() or "").strip()
        if not share_url:
            messagebox.showwarning("输入有误", "请先填写原始代理 (share_url) 再保存。")
            return False
        # 按share_url更新或追加
        found = False
        for p in self._proxies:
            if (p.get("share_url") or "").strip() == share_url:
                p["forward"] = forward
                p["proxy_port"] = proxy_port
                found = True
                break
        if not found:
            self._proxies.append({"forward": forward, "share_url": share_url, "proxy_port": proxy_port})
        save_saved_proxies(self._proxies)
        self._refresh_proxy_combo()
        self._log("已保存代理配置: " + share_url)
        return True

    def _remove_proxy(self):
        """移除选中的代理配置"""
        sel = (self.var_share_url.get() or "").strip()
        if not sel:
            messagebox.showwarning("未选择", "请先输入或选择要移除的原始代理 (share_url)。")
            return
        n_before = len(self._proxies)
        self._proxies = [p for p in self._proxies if (p.get("share_url") or "").strip() != sel]
        if len(self._proxies) >= n_before:
            messagebox.showinfo("提示", "该代理配置未在保存列表中。")
            return
        save_saved_proxies(self._proxies)
        self._refresh_proxy_combo()
        self.var_forward.set("socks5://127.0.0.1:10809")
        self.var_share_url.set("")
        self.var_proxy_port.set("10808")
        self._log("已移除代理配置: " + sel)

    def _on_email_select(self, *args):
        """邮箱下拉选已保存账号时，自动填密码和 IMAP"""
        self._try_fill_account_fields(self.var_email.get())

    def _save_current_account(self) -> bool:
        """保存当前表单中的账号到列表并持久化。已存在则更新。返回是否成功。"""
        email = (self.var_email.get() or "").strip()
        pwd = (self.var_pwd.get() or "").strip()
        imap = (self.var_imap.get() or "").strip()
        if not email or not pwd:
            messagebox.showwarning("输入有误", "请先填写邮箱和密码再保存。")
            return False
        # 按邮箱更新或追加
        found = False
        for a in self._accounts:
            if (a.get("email") or "").strip() == email:
                a["password"] = pwd
                a["imap"] = imap
                found = True
                break
        if not found:
            self._accounts.append({"email": email, "password": pwd, "imap": imap})
        save_saved_accounts(self._accounts)
        self._refresh_account_combo()
        self._log("已保存账号: " + email)
        return True

    def _build_fb_account_manager_tab(self, parent: tk.Widget):
        """在主窗口的“账号管理”页构建 UI（无弹窗）。"""
        accounts: list[dict] = load_saved_fb_accounts()

        # 变量（平台过滤 / 表单）
        var_filter_platform = tk.StringVar(value="全部")
        var_platform = tk.StringVar(value="")
        var_email = tk.StringVar()
        var_email_pwd = tk.StringVar()
        var_fb_pwd = tk.StringVar()
        var_account_id = tk.StringVar()
        var_reg_date = tk.StringVar()
        var_share_url = tk.StringVar()
        var_browser_profile = tk.StringVar()

        # 默认值：当前邮箱、当前代理、今日日期
        def _defaults() -> dict:
            p = (var_filter_platform.get() or "").strip()
            if p in ("", "全部", "未设置"):
                p = "facebook"
            return {
                "platform": p,
                "cookies": [],
                "email": (self.var_email.get() or "").strip(),
                "email_password": (self.var_pwd.get() or "").strip(),
                "password": "",
                "account_id": "",
                "regDate": _dt.date.today().isoformat(),
                "share_url": (self.var_share_url.get() or "").strip(),
                "browser_profile": (getattr(self, "_browser_var_snapshot", None).get() if hasattr(self, "_browser_var_snapshot") else "").strip()
                if hasattr(self, "_browser_var_snapshot")
                else "",
            }

        def _norm_platform(s: str) -> str:
            s = (s or "").strip()
            return s if s else ""

        def _platform_display(s: str) -> str:
            s = _norm_platform(s)
            return s if s else "未设置"

        def _get_platforms() -> list[str]:
            s = {_platform_display(a.get("platform", "")) for a in accounts}
            s.discard("")
            items = sorted(s)
            return ["全部"] + items

        def _label_of(a: dict) -> str:
            # 左侧树节点仅展示：平台 + 邮箱（不显示括号备注/ID/密码等）
            em = (a.get("email") or "").strip() or "未填写邮箱"
            plat = _platform_display(a.get("platform", ""))
            return f"【{plat}】 {em}"

        # UI
        main = ttk.Frame(parent, padding=10)
        main.pack(fill=tk.BOTH, expand=True)

        # 内容区（左右）+ 底部按钮栏（全宽）
        content = ttk.Frame(main)
        content.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        bottom_bar = ttk.Frame(main)
        bottom_bar.pack(side=tk.BOTTOM, fill=tk.X, pady=(8, 0))

        left = ttk.Frame(content, width=300)
        left.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 10))
        # 固定宽度生效（不被内容撑开/压缩）
        try:
            left.pack_propagate(False)
        except Exception:
            pass
        right = ttk.Frame(content)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 左侧：平台切换 + 日期层级树
        bar = ttk.Frame(left)
        bar.pack(fill=tk.X, pady=(0, 6))
        ttk.Label(bar, text="平台").pack(side=tk.LEFT)
        combo_platform = ttk.Combobox(bar, textvariable=var_filter_platform, width=14, state="readonly")
        combo_platform.pack(side=tk.LEFT, padx=(6, 6))

        btn_refresh = ttk.Button(bar, text="刷新", command=lambda: _reload_from_disk())
        btn_refresh.pack(side=tk.RIGHT)

        tree = ttk.Treeview(left, show="tree", selectmode="extended")
        ysb = ttk.Scrollbar(left, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=ysb.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ysb.pack(side=tk.RIGHT, fill=tk.Y)

        # 导出标记：在树里用“中间一杠”显示已导出账号
        try:
            import tkinter.font as tkfont

            base_font = ttk.Style(self.root).lookup("Treeview", "font") or "TkDefaultFont"
            exported_font = tkfont.Font(font=base_font)
            exported_font.configure(overstrike=1)
            tree.tag_configure("exported", foreground="#c00", font=exported_font)
        except Exception:
            # 字体/样式获取失败时，至少用颜色区分
            try:
                tree.tag_configure("exported", foreground="#c00")
            except Exception:
                pass

        # 右侧：账号信息表单（中文）
        form = ttk.LabelFrame(right, text="账号信息", padding=10)
        form.pack(fill=tk.BOTH, expand=True)
        form.columnconfigure(1, weight=1)

        def _row(r: int, label: str, widget: tk.Widget):
            ttk.Label(form, text=label).grid(row=r, column=0, sticky=tk.W, padx=(0, 8), pady=6)
            widget.grid(row=r, column=1, sticky=tk.EW, pady=6)

        def _password_row(r: int, label: str, var: tk.StringVar):
            ttk.Label(form, text=label).grid(row=r, column=0, sticky=tk.W, padx=(0, 8), pady=6)

            box = RoundedBox(form, height=32)
            box.grid(row=r, column=1, sticky=tk.EW, pady=6)
            box.inner.columnconfigure(0, weight=1)

            ent = tk.Entry(box.inner, textvariable=var, show="•", bd=0, highlightthickness=0, relief="flat")
            ent.grid(row=0, column=0, sticky=tk.EW)

            def _toggle():
                cur = str(ent.cget("show") or "")
                eye_open = cur != ""
                ent.configure(show="" if eye_open else "•")
                try:
                    btn.set_eye_open(not eye_open)
                except Exception:
                    pass

            btn = IconButton(box.inner, kind="eye", command=_toggle)
            btn.grid(row=0, column=1, padx=(6, 2))
            btn.set_eye_open(False)

            def _on_focus_in(_e=None):
                try:
                    box.set_focused(True)
                except Exception:
                    pass

            def _on_focus_out(_e=None):
                try:
                    box.set_focused(False)
                except Exception:
                    pass

            ent.bind("<FocusIn>", _on_focus_in)
            ent.bind("<FocusOut>", _on_focus_out)

        _row(0, "平台", ttk.Entry(form, textvariable=var_platform, width=65))
        _row(1, "邮箱", ttk.Entry(form, textvariable=var_email, width=65))
        _password_row(2, "邮箱密码", var_email_pwd)
        _password_row(3, "账号密码", var_fb_pwd)
        _row(4, "账号ID", ttk.Entry(form, textvariable=var_account_id, width=65))
        _row(5, "注册日期", ttk.Entry(form, textvariable=var_reg_date, width=65))
        _row(6, "代理", ttk.Entry(form, textvariable=var_share_url, width=65))
        combo_browser_profile = ttk.Combobox(form, textvariable=var_browser_profile, width=65)
        _row(7, "浏览器环境", combo_browser_profile)
        ttk.Label(form, text="Cookies").grid(row=8, column=0, sticky=tk.NW, padx=(0, 8), pady=6)
        txt_cookies = scrolledtext.ScrolledText(form, height=12, font=("Consolas", 10))
        txt_cookies.grid(row=8, column=1, sticky=tk.EW, pady=6)

        btns = bottom_bar  # 使用页面底部的按钮栏

        selected: dict[str, object] = {"i": None}  # 当前表单绑定的 accounts 索引（单选时）
        iid_to_index: dict[str, int] = {}

        def _set_form(d: dict):
            var_platform.set(_norm_platform(d.get("platform", "")))
            var_email.set((d.get("email") or "").strip())
            var_email_pwd.set((d.get("email_password") or "").strip())
            var_fb_pwd.set((d.get("password") or "").strip())
            var_account_id.set((d.get("account_id") or "").strip())
            var_reg_date.set((d.get("regDate") or "").strip())
            var_share_url.set((d.get("share_url") or "").strip())
            var_browser_profile.set((d.get("browser_profile") or d.get("browser_name") or "").strip())
            cookies = d.get("cookies", [])
            try:
                cookies_txt = json.dumps(cookies if isinstance(cookies, list) else [], ensure_ascii=False, indent=2)
            except Exception:
                cookies_txt = "[]"
            txt_cookies.delete("1.0", tk.END)
            txt_cookies.insert("1.0", cookies_txt)

        def _get_form() -> dict | None:
            email = (var_email.get() or "").strip()
            if not email:
                messagebox.showwarning("输入有误", "邮箱不能为空。", parent=self.root)
                return None

            cookies_raw = (txt_cookies.get("1.0", tk.END) or "").strip()

            def _guess_domain_from_hint(text: str, platform_text: str) -> str:
                # 1) 从 callback-url 推断（常见 next-auth）
                try:
                    import urllib.parse as _up

                    m = re.search(r"(?i)(?:callback-url|callback_url)=([^;]+)", text or "")
                    if m:
                        u = _up.unquote(m.group(1).strip())
                        if not u.startswith("http"):
                            u = "https://" + u.lstrip("/")
                        netloc = _up.urlparse(u).netloc.split("@")[-1].split(":")[0].strip()
                        netloc = netloc[4:] if netloc.lower().startswith("www.") else netloc
                        if netloc:
                            return "." + netloc
                except Exception:
                    pass
                # 2) 平台兜底
                p = (platform_text or "").strip().lower()
                if "facebook" in p or p == "fb":
                    return ".facebook.com"
                if p in ("x", "twitter") or "twitter" in p:
                    return ".x.com"
                if "google" in p:
                    return ".google.com"
                return ""

            def _parse_cookie_header_to_objects(text: str, platform_text: str) -> list[dict]:
                import time as _time

                s = (text or "").strip()
                if not s:
                    return []
                # 兼容换行粘贴
                s = s.replace("\r", " ").replace("\n", " ").strip()
                parts = [p.strip() for p in s.split(";") if p.strip()]
                if not parts:
                    return []

                # 可选：允许用户在最前面写 domain=.xxx.com
                domain = ""
                if parts and re.fullmatch(r"(?i)domain\s*=\s*.+", parts[0]):
                    try:
                        domain = parts[0].split("=", 1)[1].strip()
                        parts = parts[1:]
                    except Exception:
                        domain = ""
                if not domain:
                    domain = _guess_domain_from_hint(s, platform_text)

                exp = float(_time.time() + 3600 * 24 * 365 * 5)  # 默认 5 年后过期（示意）
                out: list[dict] = []
                for p0 in parts:
                    if "=" not in p0:
                        continue
                    name, value = p0.split("=", 1)
                    name = (name or "").strip()
                    value = (value or "").strip()
                    if not name:
                        continue
                    secure = name.startswith("__Secure-") or name.startswith("__Host-")
                    host_only = False
                    d = (domain or "").strip()
                    if name.startswith("__Host-"):
                        # __Host- 要求 hostOnly 且 domain 不应以点开头（尽量贴合规范）
                        host_only = True
                        if d.startswith("."):
                            d = d[1:]
                    out.append(
                        {
                            "domain": d,
                            "expirationDate": exp,
                            "hostOnly": host_only,
                            "httpOnly": False,
                            "name": name,
                            "path": "/",
                            "sameSite": "no_restriction",
                            "secure": bool(secure),
                            "session": False,
                            "storeId": None,
                            "value": value,
                        }
                    )
                return out

            cookies_val: list[dict] = []
            if not cookies_raw:
                cookies_val = []
            else:
                # 先按 JSON 解析；失败则按 Cookie header 解析
                try:
                    j = json.loads(cookies_raw)
                    if isinstance(j, list):
                        cookies_val = j
                    else:
                        messagebox.showerror("输入有误", "Cookies JSON 必须是数组（例如 [] 或 [{...}]）。", parent=self.root)
                        return None
                except Exception:
                    cookies_val = _parse_cookie_header_to_objects(cookies_raw, var_platform.get())
                    if not cookies_val:
                        messagebox.showerror(
                            "输入有误",
                            "Cookies 支持两种格式：\n"
                            "1) JSON 数组（例如 [] 或 [{...}]）\n"
                            "2) Cookie 字符串（例如 a=b; c=d; ...）",
                            parent=self.root,
                        )
                        return None

            return {
                "platform": _norm_platform(var_platform.get()),
                "cookies": cookies_val,
                "email": email,
                "email_password": (var_email_pwd.get() or "").strip(),
                "password": (var_fb_pwd.get() or "").strip(),
                "account_id": (var_account_id.get() or "").strip(),
                "regDate": (var_reg_date.get() or "").strip(),
                "share_url": (var_share_url.get() or "").strip(),
                "browser_profile": (var_browser_profile.get() or "").strip(),
            }

        def _reload_from_disk():
            accounts.clear()
            accounts.extend(load_saved_fb_accounts())
            # 浏览器环境下拉：从 saved_browser_profiles.json 读取
            try:
                bps = load_saved_browser_profiles()
                names = sorted({(p.get("name") or "").strip() for p in bps if isinstance(p, dict) and (p.get("name") or "").strip()})
                combo_browser_profile["values"] = names
            except Exception:
                pass
            combo_platform["values"] = _get_platforms()
            if var_filter_platform.get() not in combo_platform["values"]:
                var_filter_platform.set("全部")
            _refresh_tree()
            _new()

        def _sync_launch_profiles():
            """把账号+邮箱+代理+浏览器组合数据写入 saved_launch_profiles.json（用于首页一键启动）。"""
            try:
                emails_all = load_saved_accounts()
                proxies_all = load_saved_proxies()
                browsers_all = load_saved_browser_profiles()
                out: list[dict] = []
                for a in accounts:
                    if not isinstance(a, dict):
                        continue
                    out.append(
                        _build_launch_profile_for_account(
                            account=a, emails=emails_all, proxies=proxies_all, browser_profiles=browsers_all
                        )
                    )
                save_saved_launch_profiles(out)
                # 账号管理保存/删除后，首页自动刷新
                try:
                    cb = getattr(self, "_home_refresh", None)
                    if callable(cb):
                        cb()
                except Exception:
                    pass
            except Exception:
                pass

        def _refresh_tree(select_key: tuple[str, str, str] | None = None):
            iid_to_index.clear()
            tree.delete(*tree.get_children())

            fp = (var_filter_platform.get() or "").strip()
            fp_norm = "" if fp in ("全部", "未设置", "") else fp

            # 过滤 + 按 regDate 分组
            groups: dict[str, list[int]] = {}
            for idx, a in enumerate(accounts):
                plat_disp = _platform_display(a.get("platform", ""))
                if fp_norm and plat_disp != fp_norm:
                    continue
                d = (a.get("regDate") or "").strip() or "未填写日期"
                groups.setdefault(d, []).append(idx)

            # 日期倒序（YYYY-MM-DD 可直接按字符串排序），"未填写日期" 放最后
            dates = sorted([k for k in groups.keys() if k != "未填写日期"], reverse=True)
            if "未填写日期" in groups:
                dates.append("未填写日期")

            def _key_of(a: dict) -> tuple[str, str, str]:
                return (_platform_display(a.get("platform", "")), (a.get("email") or "").strip().lower(), (a.get("account_id") or "").strip())

            want_iid: str | None = None
            for date in dates:
                date_iid = f"date::{date}"
                tree.insert("", "end", iid=date_iid, text=date, open=True)
                for idx in groups[date]:
                    a = accounts[idx]
                    acc_iid = f"acc::{idx}"
                    iid_to_index[acc_iid] = idx
                    tags = ("exported",) if bool(a.get("exported")) else ()
                    tree.insert(date_iid, "end", iid=acc_iid, text=_label_of(a), tags=tags)
                    if select_key and _key_of(a) == select_key:
                        want_iid = acc_iid

            if want_iid:
                try:
                    tree.selection_set(want_iid)
                    tree.see(want_iid)
                except Exception:
                    pass

        def _selected_indices_from_tree() -> list[int]:
            sel = list(tree.selection() or [])
            out: list[int] = []
            for iid in sel:
                if iid.startswith("acc::") and iid in iid_to_index:
                    out.append(iid_to_index[iid])
                elif iid.startswith("date::"):
                    for child in tree.get_children(iid):
                        if child.startswith("acc::") and child in iid_to_index:
                            out.append(iid_to_index[child])
            # 去重 + 保序
            seen = set()
            uniq = []
            for i in out:
                if i not in seen:
                    uniq.append(i)
                    seen.add(i)
            return uniq

        def _new():
            selected["i"] = None
            tree.selection_remove(tree.selection())
            _set_form(_defaults())
            # 新建时默认跟随当前“浏览器”页选择的环境名称
            try:
                cb = getattr(self, "_accounts_set_browser_profile_default", None)
                if callable(cb):
                    cb()
            except Exception:
                pass

        def _set_browser_profile_default(force: bool = False):
            try:
                cur = (var_browser_profile.get() or "").strip()
                if cur and not force:
                    return
                snap_var = getattr(self, "_browser_var_snapshot", None)
                nm = (snap_var.get() if snap_var is not None else "").strip()
                if nm:
                    var_browser_profile.set(nm)
            except Exception:
                pass

        # 暴露给 build_ui：在浏览器页创建完后，能强制写入一次默认值
        self._accounts_set_browser_profile_default = _set_browser_profile_default

        def _save():
            item = _get_form()
            if not item:
                return

            plat_disp = _platform_display(item.get("platform", ""))
            key_email = (item.get("email") or "").strip().lower()
            key_id = (item.get("account_id") or "").strip()

            def _preserve_export_flags(old: dict, new: dict):
                # 保存/更新时保留“已导出”标记，不因编辑丢失
                for k in ("exported", "exportedAt"):
                    if k in old and k not in new:
                        new[k] = old.get(k)

            i = selected.get("i")
            if isinstance(i, int) and 0 <= i < len(accounts):
                _preserve_export_flags(accounts[i], item)
                accounts[i] = item
            else:
                # 若已存在（同平台下 account_id 或 email）则更新，否则追加
                updated = False
                for idx, a in enumerate(accounts):
                    a_plat = _platform_display(a.get("platform", ""))
                    a_email = (a.get("email") or "").strip().lower()
                    a_id = (a.get("account_id") or "").strip()
                    if a_plat != plat_disp:
                        continue
                    if (key_id and a_id and a_id == key_id) or (a_email and a_email == key_email):
                        _preserve_export_flags(accounts[idx], item)
                        accounts[idx] = item
                        updated = True
                        break
                if not updated:
                    accounts.append(item)

            save_saved_fb_accounts(accounts)
            _sync_launch_profiles()
            self._log("已保存账号信息: " + (item.get("email") or "").strip() + (f"（ID:{key_id}）" if key_id else ""))
            combo_platform["values"] = _get_platforms()
            # 保存后，尽量定位回该账号
            _refresh_tree(select_key=(plat_disp, key_email, key_id))

        def _export_selected():
            idxs = _selected_indices_from_tree()
            if not idxs:
                messagebox.showinfo("提示", "请先在左侧选择要导出的账号（可多选）。", parent=self.root)
                return
            items = [accounts[i] for i in idxs if 0 <= i < len(accounts)]
            default_name = "账号导出_所选.json"
            path = filedialog.asksaveasfilename(
                parent=self.root,
                title="导出所选账号",
                defaultextension=".json",
                initialfile=default_name,
                filetypes=[("JSON", "*.json"), ("全部", "*.*")],
            )
            if not path:
                return
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(items, f, ensure_ascii=False, indent=2)
                # 标记为“已导出”，并刷新树
                ts = _dt.datetime.now().isoformat(timespec="seconds")
                for i in idxs:
                    if 0 <= i < len(accounts):
                        accounts[i]["exported"] = True
                        accounts[i]["exportedAt"] = ts
                save_saved_fb_accounts(accounts)
                _refresh_tree()
                messagebox.showinfo("导出完成", f"已导出到：\n{path}", parent=self.root)
            except Exception as e:
                messagebox.showerror("导出失败", str(e), parent=self.root)

        def _export_all():
            if not accounts:
                messagebox.showinfo("提示", "当前没有可导出的账号。", parent=self.root)
                return
            path = filedialog.asksaveasfilename(
                parent=self.root,
                title="导出全部账号",
                defaultextension=".json",
                initialfile="账号导出_全部.json",
                filetypes=[("JSON", "*.json"), ("全部", "*.*")],
            )
            if not path:
                return
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(accounts, f, ensure_ascii=False, indent=2)
                # 标记全部为“已导出”，并刷新树
                ts = _dt.datetime.now().isoformat(timespec="seconds")
                for a in accounts:
                    if isinstance(a, dict):
                        a["exported"] = True
                        a["exportedAt"] = ts
                save_saved_fb_accounts(accounts)
                _refresh_tree()
                messagebox.showinfo("导出完成", f"已导出到：\n{path}", parent=self.root)
            except Exception as e:
                messagebox.showerror("导出失败", str(e), parent=self.root)

        def _delete():
            idxs = _selected_indices_from_tree()
            if not idxs:
                messagebox.showinfo("提示", "请先在左侧选择要删除的账号（可多选）。", parent=self.root)
                return
            if not messagebox.askyesno("确认删除", f"确定删除选中的 {len(idxs)} 个账号？\n删除后不可恢复。", parent=self.root):
                return
            for i in sorted(idxs, reverse=True):
                if 0 <= i < len(accounts):
                    accounts.pop(i)
            save_saved_fb_accounts(accounts)
            _sync_launch_profiles()
            combo_platform["values"] = _get_platforms()
            self._log(f"已删除 {len(idxs)} 个账号信息")
            _refresh_tree()
            _new()

        def _on_tree_select(_event=None):
            idxs = _selected_indices_from_tree()
            if len(idxs) == 1 and 0 <= idxs[0] < len(accounts):
                selected["i"] = idxs[0]
                _set_form(accounts[idxs[0]])
            else:
                selected["i"] = None

        tree.bind("<<TreeviewSelect>>", _on_tree_select)

        def _on_platform_changed(_event=None):
            _refresh_tree()
            _new()

        combo_platform.bind("<<ComboboxSelected>>", _on_platform_changed)

        ttk.Button(btns, text="新建", command=_new).pack(side=tk.LEFT)
        ttk.Button(btns, text="保存/更新", command=_save).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(btns, text="删除", command=_delete).pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(btns, text="导出所选", command=_export_selected).pack(side=tk.RIGHT)
        ttk.Button(btns, text="导出全部", command=_export_all).pack(side=tk.RIGHT, padx=(0, 8))

        # 初始化
        combo_platform["values"] = _get_platforms()
        _reload_from_disk()
        _refresh_tree()
        _new()

    def _build_browser_settings_tab(self, parent: tk.Widget):
        """主窗口的“浏览器”页：生成/保存/启动命令。"""

        # UA 会根据 kernel/chrome-version 自动联动填充（仅保留主版本号）
        UA_PRESETS = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36",
        ]
        # WebGL 下拉会优先从 webgl_info.json 读取，这里是兜底默认值
        VENDOR_PRESETS = ["Google Inc. (NVIDIA)", "Google Inc. (Intel)", "Google Inc. (AMD)"]
        RENDERER_PRESETS = [
            "ANGLE (NVIDIA, NVIDIA GeForce GTX 710 (0x0000128B) Direct3D11 vs_5_0 ps_5_0, D3D11)",
            "ANGLE (NVIDIA, NVIDIA GeForce RTX 4060 Laptop GPU (0x000028E0) Direct3D11 vs_5_0 ps_5_0, D3D11)",
            "ANGLE (Intel, Intel(R) UHD Graphics 770 Direct3D11 vs_5_0 ps_5_0, D3D11)",
        ]
        LANG_PRESETS = ["en-US", "zh-CN", "zh-TW", "id-ID", "vi-VN"]
        TIMEZONE_PRESETS = [
            "America/New_York",
            "America/Los_Angeles",
            "America/Chicago",
            "Europe/London",
            "Asia/Singapore",
            "Asia/Shanghai",
        ]
        CORE_PRESETS = ["2", "4", "6", "8", "12", "16", "20", "24", "32"]
        WIN_PRESETS = ["win10", "win11"]
        # chrome-version 的下拉会从 useragent_info.json 读取并按 kernel_version 联动
        CHROME_VER_PRESETS = ["142.0.7444.176", "143.0.7499.193"]

        OTHER_DEFAULT = (
            '--proxy-server="socks5://127.0.0.1:50681"  '
            "--enable-test  "
            "--disk-cache-size=1073741824 "
            '--proxy-bypass-list="localhost,127.0.0.1;172.16.*,172.17.*,172.18.*,192.168.*,10.1.1.1/8"  '
            "-disable-features=IsolateOrigins,SkiaGraphite,site-per-process "
            "--no-default-browser-check"
        )

        def _norm(s: str) -> str:
            return (s or "").strip()

        def _cmdline(args: list[str]) -> str:
            try:
                return subprocess.list2cmdline(args)
            except Exception:
                return " ".join(args)

        def _split_args(s: str) -> list[str]:
            s = _norm(s)
            if not s:
                return []
            try:
                return shlex.split(s, posix=False)
            except Exception:
                return [s]

        def _load_useragent_info() -> tuple[list[str], dict[str, list[str]]]:
            """读取同目录 useragent_info.json，返回 kernel_version 列表与 kernel->chrome_versions 映射。"""
            try:
                p = os.path.join(_app_dir(), "useragent_info.json")
                if not os.path.isfile(p):
                    return ([], {})
                with open(p, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if not isinstance(raw, dict):
                    return ([], {})

                kernel_to_versions: dict[str, list[str]] = {}

                def _ver_key(v: str) -> tuple[int, int, int, int]:
                    parts = (v or "").strip().split(".")
                    nums = []
                    for s in parts[:4]:
                        try:
                            nums.append(int(s))
                        except Exception:
                            nums.append(0)
                    while len(nums) < 4:
                        nums.append(0)
                    return tuple(nums)  # type: ignore[return-value]

                for k, arr in raw.items():
                    k = str(k).strip()
                    if not k or not isinstance(arr, list):
                        continue
                    seen = set()
                    versions: list[str] = []
                    for it in arr:
                        if not isinstance(it, dict):
                            continue
                        plat = str(it.get("platform") or "").strip().lower()
                        if plat and plat != "windows":
                            continue
                        v = str(it.get("ua_full_version") or "").strip()
                        if not v:
                            # fallback：尝试从 ua 里提取 Chrome/x.y.z.w
                            ua = str(it.get("ua") or "")
                            m = re.search(r"Chrome/([0-9]+(?:\.[0-9]+){1,3})", ua)
                            v = (m.group(1) if m else "").strip()
                        if not v or v in seen:
                            continue
                        seen.add(v)
                        versions.append(v)
                    if versions:
                        # 版本倒序（新版本在前）
                        versions = sorted(versions, key=_ver_key, reverse=True)
                        kernel_to_versions[k] = versions

                kernels = sorted(kernel_to_versions.keys(), key=lambda s: int(re.sub(r"\D", "", s) or "0"), reverse=True)
                return (kernels, kernel_to_versions)
            except Exception:
                return ([], {})

        KERNEL_PRESETS, KERNEL_TO_CHROME_VERS = _load_useragent_info()
        if KERNEL_PRESETS:
            # 默认 chrome-version 预选取第一个 kernel 的列表
            CHROME_VER_PRESETS = KERNEL_TO_CHROME_VERS.get(KERNEL_PRESETS[0], CHROME_VER_PRESETS) or CHROME_VER_PRESETS

        def _load_webgl_presets() -> tuple[list[str], dict[str, list[str]]]:
            """读取同目录 webgl_info.json，返回 vendor 列表与 vendor->renderers 映射（仅 Windows）。"""
            try:
                p = os.path.join(_app_dir(), "webgl_info.json")
                if not os.path.isfile(p):
                    return ([], {})
                with open(p, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if not isinstance(raw, dict):
                    return ([], {})
                vendor_to_renderers: dict[str, list[str]] = {}
                for vendor, arr in raw.items():
                    if not isinstance(arr, list):
                        continue
                    rset: list[str] = []
                    seen = set()
                    for it in arr:
                        if not isinstance(it, dict):
                            continue
                        sysv = (it.get("system") or "").strip()
                        if sysv and sysv.lower() != "windows":
                            continue
                        r = (it.get("unmasked_renderer") or "").strip()
                        if not r:
                            continue
                        if r not in seen:
                            rset.append(r)
                            seen.add(r)
                    if rset:
                        vendor_to_renderers[(vendor or "").strip()] = rset
                vendors = sorted([v for v in vendor_to_renderers.keys() if v])
                # 常用项置前（若存在）
                prefer = ["Google Inc. (NVIDIA)", "Google Inc. (Intel)", "Google Inc. (AMD)"]
                ordered: list[str] = []
                for pv in prefer:
                    if pv in vendor_to_renderers and pv not in ordered:
                        ordered.append(pv)
                for v in vendors:
                    if v not in ordered:
                        ordered.append(v)
                return (ordered, vendor_to_renderers)
            except Exception:
                return ([], {})

        webgl_vendors, webgl_vendor_to_renderers = _load_webgl_presets()
        if webgl_vendors:
            VENDOR_PRESETS = webgl_vendors
            # 默认 renderer 用第一个 vendor 的第一个 renderer
            first_r = webgl_vendor_to_renderers.get(webgl_vendors[0]) or []
            if first_r:
                RENDERER_PRESETS = first_r

        profiles: list[dict] = load_saved_browser_profiles()
        # 暴露给“首页/账号管理”使用（引用同一个 list，保存时会原地更新）
        self._browser_profiles = profiles

        def _current_username() -> str:
            return (os.environ.get("USERNAME") or os.environ.get("USER") or "admin").strip() or "admin"

        _user = _current_username()
        DEFAULT_EXE_PATH = rf"C:\Users\{_user}\AppData\Local\Chromium\Application\chrome.exe"
        DEFAULT_USERDIR = rf"C:\Users\{_user}\AppData\Local\Chromium\User Data"
        # 默认插件（放在程序同目录即可；注意 Chrome 需要解包目录，.crx 会在启动时自动解包）
        DEFAULT_CRX_PATH = os.path.join(_app_dir(), "cookie_editor_v2.2.0.0.crx")

        def _rand_digits_9_10() -> str:
            """生成 9~10 位数字字符串（用于 fp / fp-font 默认值）。"""
            import secrets

            n = 9 + secrets.randbelow(2)  # 9 or 10
            base = 10 ** (n - 1)
            return str(base + secrets.randbelow(9 * base))

        # --- 顶部：快照 ---
        wrap = ttk.Frame(parent, padding=8)
        wrap.pack(fill=tk.BOTH, expand=True)
        wrap.columnconfigure(1, weight=1)

        top = ttk.Frame(wrap)
        top.grid(row=0, column=0, columnspan=2, sticky=tk.EW, pady=(0, 8))
        top.columnconfigure(1, weight=1)

        var_snapshot = tk.StringVar()
        # 暴露当前“环境名称”
        self._browser_var_snapshot = var_snapshot
        combo_snapshot = ttk.Combobox(top, textvariable=var_snapshot, width=44)

        ttk.Label(top, text="名称:", width=12, anchor="w").grid(row=0, column=0, sticky=tk.W)
        combo_snapshot.grid(row=0, column=1, sticky=tk.EW, padx=(6, 6))

        def _delete_current():
            name = _norm(var_snapshot.get())
            if not name:
                messagebox.showinfo("提示", "请先选择一个快照。", parent=self.root)
                return
            idx = next((i for i, p in enumerate(profiles) if _norm(p.get("name", "")) == name), None)
            if idx is None:
                messagebox.showinfo("提示", "当前快照未找到。", parent=self.root)
                return
            if not messagebox.askyesno("确认", f"确定删除快照：{name}？", parent=self.root):
                return
            profiles.pop(idx)
            save_saved_browser_profiles(profiles)
            self._log("已删除快照: " + name)
            _refresh_snapshot_list()
            _clear_form()

        ttk.Button(top, text="删除", command=_delete_current).grid(row=0, column=2, sticky=tk.W)

        # --- 表单变量 ---
        var_exe = tk.StringVar(value=DEFAULT_EXE_PATH)
        # kernel_version：用于联动 chrome-version / useragent
        var_kernel = tk.StringVar(value=(KERNEL_PRESETS[0] if 'KERNEL_PRESETS' in locals() and KERNEL_PRESETS else "142"))
        # fp / fp-font：默认打开软件自动生成 9~10 位随机数字
        var_fp = tk.StringVar(value=_rand_digits_9_10())
        var_fp_font = tk.StringVar(value=_rand_digits_9_10())
        var_lang = tk.StringVar(value="en-US")
        var_core = tk.StringVar(value="8")
        var_vendor = tk.StringVar(value=VENDOR_PRESETS[0] if VENDOR_PRESETS else "")
        var_renderer = tk.StringVar(value=RENDERER_PRESETS[0] if RENDERER_PRESETS else "")
        var_ua = tk.StringVar(value=UA_PRESETS[0] if UA_PRESETS else "")
        var_winver = tk.StringVar(value="win11")
        var_chromever = tk.StringVar(value=CHROME_VER_PRESETS[0] if CHROME_VER_PRESETS else "")
        var_webrtc = tk.StringVar(value="")
        var_userdir = tk.StringVar(value=DEFAULT_USERDIR)
        var_timezone = tk.StringVar(value="America/New_York")
        # 默认不加载插件（需要可在下拉/浏览中选择）
        var_crx = tk.StringVar(value="")

        # 供“测试代理”联动回填
        self.var_browser_webrtc = var_webrtc
        self.var_browser_timezone = var_timezone
        # 如果之前已测试到 IP/时区，则在进入浏览器页时自动带上
        try:
            ip0 = getattr(self, "_last_proxy_test_ip", "") or ""
            tz0 = getattr(self, "_last_proxy_test_timezone", "") or ""
            if ip0:
                self.var_browser_webrtc.set(ip0)
            if tz0:
                self.var_browser_timezone.set(tz0)
        except Exception:
            pass

        # --- exe 路径行 ---
        exe_row = ttk.Frame(wrap)
        exe_row.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(0, 8))
        exe_row.columnconfigure(1, weight=1)
        ttk.Label(exe_row, text="浏览器路径:", width=12, anchor="w").grid(row=0, column=0, sticky=tk.W)
        combo_exe = ttk.Combobox(exe_row, textvariable=var_exe)
        combo_exe.grid(row=0, column=1, sticky=tk.EW, padx=(6, 6))

        def _pick_exe():
            pth = filedialog.askopenfilename(parent=self.root, title="选择浏览器 exe", filetypes=[("EXE", "*.exe"), ("全部", "*.*")])
            if pth:
                var_exe.set(pth)
                _update_preview()

        ttk.Button(exe_row, text="浏览…", command=_pick_exe).grid(row=0, column=2, sticky=tk.W)

        # --- 浏览器目录（user-data-dir）放在浏览器路径下方 ---
        userdir_row = ttk.Frame(wrap)
        userdir_row.grid(row=2, column=0, columnspan=2, sticky=tk.EW, pady=(0, 8))
        userdir_row.columnconfigure(1, weight=1)
        ttk.Label(userdir_row, text="浏览器目录:", width=12, anchor="w").grid(row=0, column=0, sticky=tk.W)
        combo_userdir = ttk.Combobox(userdir_row, textvariable=var_userdir, values=[], width=52)
        combo_userdir.grid(row=0, column=1, sticky=tk.EW, padx=(6, 6))

        def _pick_userdir():
            pth = filedialog.askdirectory(parent=self.root, title="选择浏览器目录（user-data-dir）")
            if pth:
                var_userdir.set(pth)

        ttk.Button(userdir_row, text="浏览…", command=_pick_userdir).grid(row=0, column=2, sticky=tk.W)

        # --- 参数区域（两列） ---
        grid = ttk.Frame(wrap)
        # 参数区不再纵向撑满（避免 useragent 与 others 之间出现大片空白）
        grid.grid(row=3, column=0, columnspan=2, sticky=tk.EW)
        wrap.rowconfigure(3, weight=0)
        grid.columnconfigure(1, weight=1)
        grid.columnconfigure(4, weight=1)

        def _row(r: int, left_label: str, left_widget: tk.Widget, right_label: str, right_widget: tk.Widget):
            """两列参数行（去掉随机按钮后更紧凑）。"""
            ttk.Label(grid, text=left_label).grid(row=r, column=0, sticky=tk.W, pady=4)
            left_widget.grid(row=r, column=1, sticky=tk.EW, pady=4, padx=(6, 6))
            ttk.Label(grid, text=right_label).grid(row=r, column=3, sticky=tk.W, pady=4, padx=(10, 0))
            right_widget.grid(row=r, column=4, sticky=tk.EW, pady=4, padx=(6, 6))

        combo_fp = ttk.Combobox(grid, textvariable=var_fp, values=[], width=28)
        combo_fp_font = ttk.Combobox(grid, textvariable=var_fp_font, values=[], width=28)
        combo_webrtc = ttk.Combobox(grid, textvariable=var_webrtc, values=[], width=28)
        combo_kernel = ttk.Combobox(
            grid,
            textvariable=var_kernel,
            values=(KERNEL_PRESETS if 'KERNEL_PRESETS' in locals() else []),
            width=28,
            state="readonly" if ('KERNEL_PRESETS' in locals() and KERNEL_PRESETS) else "normal",
        )
        combo_cver = ttk.Combobox(grid, textvariable=var_chromever, values=CHROME_VER_PRESETS, width=28)
        combo_win = ttk.Combobox(grid, textvariable=var_winver, values=WIN_PRESETS, width=28)
        combo_core = ttk.Combobox(grid, textvariable=var_core, values=CORE_PRESETS, width=28)
        combo_lang = ttk.Combobox(grid, textvariable=var_lang, values=LANG_PRESETS, width=28)
        combo_tz = ttk.Combobox(grid, textvariable=var_timezone, values=TIMEZONE_PRESETS, width=28)
        combo_vendor = ttk.Combobox(grid, textvariable=var_vendor, values=VENDOR_PRESETS, width=28)
        combo_renderer = ttk.Combobox(grid, textvariable=var_renderer, values=RENDERER_PRESETS)
        combo_ua = ttk.Combobox(grid, textvariable=var_ua, values=UA_PRESETS)

        def _sync_renderer_list(_event=None):
            v = _norm(var_vendor.get())
            rs = webgl_vendor_to_renderers.get(v) or RENDERER_PRESETS or []
            try:
                combo_renderer["values"] = rs
            except Exception:
                pass
            cur = _norm(var_renderer.get())
            if rs and cur not in rs:
                var_renderer.set(rs[0])
            elif not rs and cur:
                # 没有可选项时保留当前输入
                pass

        combo_vendor.bind("<<ComboboxSelected>>", _sync_renderer_list)
        try:
            var_vendor.trace_add("write", lambda *_a: _sync_renderer_list())
        except Exception:
            pass

        def _major_of_chrome_ver(s: str) -> str:
            s = _norm(s)
            if not s:
                return ""
            out = []
            for ch in s:
                if ch.isdigit():
                    out.append(ch)
                elif ch == ".":
                    break
                else:
                    break
            return "".join(out)

        def _ua_for_major(major: str) -> str:
            major = _norm(major) or "142"
            # 只保留主版本号，不使用具体补丁版本
            return (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                f"Chrome/{major}.0.0.0 Safari/537.36"
            )

        def _sync_chrome_versions_for_kernel(_event=None):
            k = _norm(var_kernel.get())
            m: dict[str, list[str]] = KERNEL_TO_CHROME_VERS if isinstance(KERNEL_TO_CHROME_VERS, dict) else {}
            cvs = m.get(k) or []
            if not cvs:
                cvs = CHROME_VER_PRESETS
            try:
                combo_cver["values"] = cvs
            except Exception:
                pass
            cur = _norm(var_chromever.get())
            if cvs and cur not in cvs:
                var_chromever.set(cvs[0])

        def _sync_ua_from_kernel_or_chrome(_event=None):
            major = _norm(var_kernel.get()) or _major_of_chrome_ver(var_chromever.get()) or "142"
            ua = _ua_for_major(major)
            try:
                combo_ua["values"] = [ua]
            except Exception:
                pass
            var_ua.set(ua)

        combo_kernel.bind("<<ComboboxSelected>>", lambda _e: (_sync_chrome_versions_for_kernel(), _sync_ua_from_kernel_or_chrome()))
        combo_cver.bind("<<ComboboxSelected>>", _sync_ua_from_kernel_or_chrome)
        try:
            var_kernel.trace_add("write", lambda *_a: (_sync_chrome_versions_for_kernel(), _sync_ua_from_kernel_or_chrome()))
        except Exception:
            pass
        try:
            var_chromever.trace_add("write", lambda *_a: _sync_ua_from_kernel_or_chrome())
        except Exception:
            pass

        _row(0, "fp:", combo_fp, "fp-font:", combo_fp_font)
        _row(1, "webrtc:", combo_webrtc, "（保留）", ttk.Label(grid, text=""))

        # extension-crx（放在 webrtc 后面；默认无值；支持下拉选择）
        ttk.Label(grid, text="extension-crx:").grid(row=2, column=0, sticky=tk.W, pady=4)
        ext_row = ttk.Frame(grid)
        ext_row.grid(row=2, column=1, columnspan=4, sticky=tk.EW, pady=4, padx=(6, 6))
        ext_row.columnconfigure(0, weight=1)
        combo_crx = ttk.Combobox(ext_row, textvariable=var_crx, values=[], width=28)
        combo_crx.grid(row=0, column=0, sticky=tk.EW, padx=(0, 6))

        def _pick_crx():
            pth = filedialog.askopenfilename(
                parent=self.root,
                title="选择扩展（.crx 或已解包目录）",
                filetypes=[("CRX", "*.crx"), ("全部", "*.*")],
            )
            if pth:
                var_crx.set(pth)

        ttk.Button(ext_row, text="浏览…", command=_pick_crx).grid(row=0, column=1, sticky=tk.W)

        _row(3, "kernel_version:", combo_kernel, "chrome-version:", combo_cver)
        _row(4, "win-version:", combo_win, "cores:", combo_core)
        _row(5, "language:", combo_lang, "timezone:", combo_tz)

        # webgl-vendor (span)
        ttk.Label(grid, text="webgl-vendor:").grid(row=6, column=0, sticky=tk.W, pady=4)
        combo_vendor.grid(row=6, column=1, columnspan=4, sticky=tk.EW, pady=4, padx=(6, 6))

        # webgl-renderer (span)
        ttk.Label(grid, text="webgl-renderer:").grid(row=7, column=0, sticky=tk.W, pady=4)
        combo_renderer.grid(row=7, column=1, columnspan=4, sticky=tk.EW, pady=4, padx=(6, 6))

        # useragent (span)
        # useragent 与下方 others 的间距保持与上方参数行一致
        ttk.Label(grid, text="useragent:").grid(row=8, column=0, sticky=tk.W, pady=4)
        combo_ua.grid(row=8, column=1, columnspan=4, sticky=tk.EW, pady=4, padx=(6, 6))

        # --- others + 结果（上下布局，随窗口伸缩） ---
        bottom = ttk.Frame(wrap)
        # 更贴近上方 useragent 行
        bottom.grid(row=4, column=0, columnspan=2, sticky=tk.NSEW, pady=(4, 6))
        bottom.columnconfigure(0, weight=1)
        # bottom 两个框尽量铺满窗口可用空间
        bottom.rowconfigure(1, weight=2)
        # 结果预览更高一些（随窗口伸缩时占比更大）
        bottom.rowconfigure(3, weight=5)
        wrap.rowconfigure(4, weight=6)

        ttk.Label(bottom, text="others:").grid(row=0, column=0, sticky=tk.W)
        txt_other = scrolledtext.ScrolledText(bottom, height=4, font=("Consolas", 10))
        txt_other.grid(row=1, column=0, sticky=tk.NSEW, pady=(2, 4))
        txt_other.insert("1.0", OTHER_DEFAULT)
        # 暴露 others 输入框，便于首页一键启动时同步代理端口
        self._browser_txt_other = txt_other

        ttk.Label(bottom, text="结果预览:").grid(row=2, column=0, sticky=tk.W)
        # 结果预览默认再高 ~50%
        txt_cmd = scrolledtext.ScrolledText(bottom, height=8, font=("Consolas", 9))
        txt_cmd.grid(row=3, column=0, sticky=tk.NSEW, pady=(2, 0))
        self._browser_txt_cmd = txt_cmd

        # --- 底部按钮 ---
        btns = ttk.Frame(wrap)
        # 按钮与上方底部区域稍微拉开
        btns.grid(row=5, column=0, columnspan=2, sticky=tk.EW, pady=(10, 0))
        btns.columnconfigure(0, weight=1)

        def _clear_form():
            var_fp.set(_rand_digits_9_10())
            var_fp_font.set(_rand_digits_9_10())
            var_lang.set("en-US")
            var_core.set("8")
            var_vendor.set(VENDOR_PRESETS[0] if VENDOR_PRESETS else "")
            # vendor 变化时会联动刷新 renderer 列表，这里先清空再交给联动
            var_renderer.set("")
            var_winver.set("win11")
            # kernel/chrome-version/ua 会联动刷新
            try:
                var_kernel.set((KERNEL_PRESETS[0] if KERNEL_PRESETS else "142"))
            except Exception:
                var_kernel.set("142")
            var_chromever.set(CHROME_VER_PRESETS[0] if CHROME_VER_PRESETS else "")
            var_webrtc.set("")
            # 清理后恢复默认浏览器目录
            var_userdir.set(DEFAULT_USERDIR)
            var_crx.set("")
            var_timezone.set("America/New_York")
            txt_other.delete("1.0", tk.END)
            txt_other.insert("1.0", OTHER_DEFAULT)
            txt_cmd.delete("1.0", tk.END)
            try:
                _sync_chrome_versions_for_kernel()
                _sync_ua_from_kernel_or_chrome()
            except Exception:
                pass

        def _profile_from_form(name: str | None = None) -> dict:
            # 额外保存“原始代理”（收件箱里的 share_url），方便选择环境后自动回填
            try:
                share_url = _norm(getattr(self, "var_share_url", None).get()) if hasattr(self, "var_share_url") else ""
            except Exception:
                share_url = ""
            # 如果收件箱里当前没填原始代理，则用“名称”兜底（很多情况下名称就是环境/代理标识）
            share_url = share_url or _norm(name or var_snapshot.get())
            return {
                "name": _norm(name or var_snapshot.get()),
                "exe_path": _norm(var_exe.get()),
                "share_url": share_url,
                "kernel_version": _norm(var_kernel.get()),
                "fp": _norm(var_fp.get()),
                "fp_font": _norm(var_fp_font.get()),
                "language": _norm(var_lang.get()),
                "core": _norm(var_core.get()),
                "vendor": _norm(var_vendor.get()),
                "renderer": _norm(var_renderer.get()),
                "ua": _norm(var_ua.get()),
                "win_version": _norm(var_winver.get()),
                "chrome_version": _norm(var_chromever.get()),
                "webrtc": _norm(var_webrtc.get()),
                "timezone": _norm(var_timezone.get()),
                # 兜底：确保启动时总会带上 user-data-dir
                "user_dir": _norm(var_userdir.get()) or DEFAULT_USERDIR,
                "crx_path": _norm(var_crx.get()),
                "other": txt_other.get("1.0", tk.END).strip(),
            }

        _applying_snapshot = False

        def _apply_profile(p: dict):
            nonlocal _applying_snapshot
            if _applying_snapshot:
                return
            _applying_snapshot = True
            fp_raw = p.get("fp")
            fp_dict = fp_raw if isinstance(fp_raw, dict) else {}
            var_snapshot.set(_norm(p.get("name", "")))
            var_exe.set(_norm(p.get("exe_path", p.get("browser_path", ""))))
            # 选择环境时：把保存的“原始代理”回填到收件箱
            try:
                # 兼容：如果 share_url 为空，则用 name 兜底
                su = _norm(p.get("share_url", "")) or _norm(p.get("name", ""))
                if su and hasattr(self, "var_share_url"):
                    self.var_share_url.set(su)
                    # 联动填充转发地址/本地端口
                    try:
                        self._on_proxy_select()
                    except Exception:
                        pass
            except Exception:
                pass
            # 先设置 kernel，再刷新 chrome-version 下拉
            kv = _norm(p.get("kernel_version", "")) or _major_of_chrome_ver(_norm(p.get("chrome_version", "")))
            if kv:
                var_kernel.set(kv)
            try:
                _sync_chrome_versions_for_kernel()
            except Exception:
                pass
            var_fp.set(_norm(p.get("fp", "")) or _norm(fp_dict.get("fp-webgl-img", "")))
            var_fp_font.set(_norm(p.get("fp_font", "")) or _norm(fp_dict.get("fp-font", "")) or _norm(var_fp_font.get()))
            var_lang.set(_norm(p.get("language", "")) or _norm(fp_dict.get("fp-language", "")) or "en-US")
            var_core.set(_norm(p.get("core", "")) or _norm(fp_dict.get("fp-cores", "")) or "8")
            var_winver.set(_norm(p.get("win_version", "")) or _norm(fp_dict.get("fp-win-version", "")) or "win11")
            var_chromever.set(_norm(p.get("chrome_version", "")) or _norm(fp_dict.get("fp-chrome-version", "")) or (CHROME_VER_PRESETS[0] if CHROME_VER_PRESETS else ""))
            var_webrtc.set(_norm(p.get("webrtc", "")) or _norm(fp_dict.get("fp-webrtc", "")))
            var_timezone.set(_norm(p.get("timezone", "")) or _norm(fp_dict.get("fp-timezone", "")) or "America/New_York")
            # user_dir：缺省时回退到默认目录（避免显示为空）
            var_userdir.set(_norm(p.get("user_dir", "")) or _norm(p.get("user_data_dir", "")) or DEFAULT_USERDIR)
            # crx：默认不加载，只有显式配置时才加载
            var_crx.set(_norm(p.get("crx_path", "")))
            # vendor/renderer：先设 vendor，再根据 vendor 刷新 renderer 可选项
            var_vendor.set(_norm(p.get("vendor", "")) or _norm(fp_dict.get("fp-webgl-vendor", "")) or (VENDOR_PRESETS[0] if VENDOR_PRESETS else ""))
            var_renderer.set(_norm(p.get("renderer", "")) or _norm(fp_dict.get("fp-webgl-renderer", "")) or (RENDERER_PRESETS[0] if RENDERER_PRESETS else ""))
            _sync_renderer_list()
            try:
                _sync_ua_from_kernel_or_chrome()
            except Exception:
                pass
            txt_other.delete("1.0", tk.END)
            txt_other.insert("1.0", _norm(p.get("other", "")) or _norm(p.get("extra_args", "")) or OTHER_DEFAULT)
            _update_preview()
            _applying_snapshot = False

        def _build_args(p: dict) -> list[str]:
            exe = _norm(p.get("exe_path", ""))
            if not exe:
                return []
            args: list[str] = [exe]

            def _strip_wrap_quotes(v: str) -> str:
                v = (v or "").strip()
                if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
                    return v[1:-1]
                return v

            def _is_chinese(s: str) -> bool:
                return bool(re.search(r"[\u4e00-\u9fff]", s or ""))

            def _safe_leaf(s: str) -> str:
                """把名称转成可用作路径叶子节点的字符串（尽量少改）。"""
                s = (s or "").strip()
                # Windows 文件名非法字符：<>:"/\|?*
                bad = '<>:"/\\|?*'
                for ch in bad:
                    s = s.replace(ch, "_")
                s = re.sub(r"\s+", "_", s).strip("_")
                return s[:64] if len(s) > 64 else s

            user_dir_base = _strip_wrap_quotes(_norm(p.get("user_dir", "")))
            if user_dir_base:
                nm = _norm(p.get("name", "")) or _norm(var_snapshot.get())
                fpv = _norm(p.get("fp", "")) or _norm(var_fp.get())
                suffix = ""
                if nm:
                    if _is_chinese(nm):
                        suffix = fpv
                    else:
                        # 英文数字/常见连接符：直接用名称
                        if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{0,63}", nm):
                            suffix = nm
                        else:
                            suffix = _safe_leaf(nm)
                else:
                    # 名称为空：也拼接 fp
                    suffix = fpv
                # “浏览器目录再拼接一层”
                eff = user_dir_base
                if suffix:
                    try:
                        tail = os.path.basename(os.path.normpath(user_dir_base))
                        if tail != suffix:
                            eff = os.path.join(user_dir_base, suffix)
                    except Exception:
                        eff = os.path.join(user_dir_base, suffix)
                args.append(f'--user-data-dir={eff}')

            def _ensure_unpacked_extension_dir(path_or_crx: str) -> str:
                """把 .crx 解包成扩展目录并返回目录路径；若传入已解包目录则直接返回。"""
                pth = (path_or_crx or "").strip().strip('"').strip("'")
                if not pth:
                    return ""
                # 允许直接传入目录（manifest.json）
                try:
                    if os.path.isdir(pth) and os.path.isfile(os.path.join(pth, "manifest.json")):
                        return pth
                except Exception:
                    pass
                if not (os.path.isfile(pth) and pth.lower().endswith(".crx")):
                    return ""
                base = os.path.splitext(os.path.basename(pth))[0]
                out_dir = os.path.join(_app_dir(), "unpacked_extensions", _safe_leaf(base or "ext"))
                manifest = os.path.join(out_dir, "manifest.json")
                try:
                    if os.path.isfile(manifest):
                        return out_dir
                    os.makedirs(out_dir, exist_ok=True)
                    with open(pth, "rb") as f:
                        data = f.read()
                    idx = data.find(b"PK\x03\x04")
                    if idx < 0:
                        return ""
                    import io
                    import zipfile

                    with zipfile.ZipFile(io.BytesIO(data[idx:])) as zf:
                        zf.extractall(out_dir)
                    return out_dir if os.path.isfile(manifest) else out_dir
                except Exception:
                    return ""

            # extension：默认加载 cookie_editor_v2.2.0.0.crx（若存在）；也可手动选择其它 crx/目录
            crx_path = _norm(p.get("crx_path", "")) or (DEFAULT_CRX_PATH if os.path.isfile(DEFAULT_CRX_PATH) else "")
            ext_dir = _ensure_unpacked_extension_dir(crx_path)
            if ext_dir:
                args.append(f'--disable-extensions-except={ext_dir}')
                args.append(f'--load-extension={ext_dir}')

            def _sanitize_token(tok: str) -> str:
                tok = (tok or "").strip()
                if not tok or "=" not in tok:
                    return tok
                k, v = tok.split("=", 1)
                kl = (k or "").strip().lower()
                # 代理参数：实际启动时不要把值包在引号里，否则 Windows 会把引号转义成 \"...\"（你截图里的问题）
                if kl in ("--proxy-server", "--proxy-bypass-list"):
                    v = v.replace('\\"', '"')
                    v = _strip_wrap_quotes(v)
                    return f"{k}={v}"
                return tok

            # other：把 disable-features/no-default 放到后面（更接近你给的示例顺序）
            other_raw = _norm(p.get("other", ""))
            other_tokens = _split_args(other_raw)
            end_tokens: list[str] = []
            pre_tokens: list[str] = []
            for t in other_tokens:
                tl = t.strip().lower()
                if tl.startswith("-disable-features=") or tl.startswith("--disable-features=") or tl == "--no-default-browser-check":
                    end_tokens.append(t)
                else:
                    pre_tokens.append(t)
            args.extend([_sanitize_token(t) for t in pre_tokens if (t or "").strip()])

            vendor = _norm(p.get("vendor", ""))
            if vendor:
                args.append(f'--fp-webgl-vendor={vendor}')
            renderer = _norm(p.get("renderer", ""))
            if renderer:
                args.append(f'--fp-webgl-renderer={renderer}')

            core = _norm(p.get("core", ""))
            if core:
                args.append(f'--fp-cores={core}')
            wv = _norm(p.get("win_version", ""))
            if wv:
                args.append(f'--fp-win-version={wv}')

            args.extend([_sanitize_token(t) for t in end_tokens if (t or "").strip()])

            tz = _norm(p.get("timezone", ""))
            if tz:
                # args.append(f'--time-zone-for-testing={tz}')
                args.append(f'--fp-timezone={tz}')

            fp_font = _norm(p.get("fp_font", ""))
            if fp_font:
                args.append(f'--fp-font={fp_font}')

            lang = _norm(p.get("language", ""))
            if lang:
                args.append(f'--lang={lang}')
                args.append(f'--fp-language={lang}')

            ua = _norm(p.get("ua", ""))
            if ua:
                args.append(f'--fp-useragent={ua}')

            cv = _norm(p.get("chrome_version", ""))
            if cv:
                args.append(f'--fp-chrome-version={cv}')

            fp = _norm(p.get("fp", ""))
            if fp:
                args.append(f'--fp-audio={fp}')
                args.append(f'--fp-webgl-img={fp}')
                args.append(f'--fp-canvas={fp}')

            webrtc = _norm(p.get("webrtc", ""))
            if webrtc:
                args.append(f'--fp-webrtc={webrtc}')

            return args

        def _args_to_preview(args: list[str]) -> str:
            """把参数转成你想要的展示格式：exe/user-data-dir/fp-* 等自动加引号。"""

            def _q(v: str) -> str:
                return '"' + (v or "").replace('"', '\\"') + '"'

            def _strip_wrap_quotes(v: str) -> str:
                v = (v or "").strip()
                if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
                    return v[1:-1]
                return v

            def _render_one(a: str, is_first: bool) -> str:
                a = (a or "").strip()
                if not a:
                    return ""
                if is_first:
                    return _q(a)
                if "=" not in a:
                    return a
                k, v = a.split("=", 1)
                kl = k.lower()
                # 这些值在示例里是带引号的
                if kl in ("--fp-timezone",):
                    return a
                if kl in ("--user-data-dir", "--proxy-server", "--proxy-bypass-list", "--load-extension", "--disable-extensions-except") or kl.startswith("--fp-"):
                    # v 可能已经带了引号（来自 others），这里先去掉外层引号，避免二次转义成 \"...\"
                    return f"{k}={_q(_strip_wrap_quotes(v))}"
                return a

            out: list[str] = []
            for i, a in enumerate(args):
                s = _render_one(a, i == 0)
                if s:
                    out.append(s)
            return "  ".join(out)

        def _update_preview():
            p = _profile_from_form()
            args = _build_args(p)
            txt_cmd.delete("1.0", tk.END)
            if not args:
                txt_cmd.insert("1.0", "（请先填写 exe 路径）")
            else:
                txt_cmd.insert("1.0", _args_to_preview(args))
        # 暴露预览刷新（首页一键启动会调用）
        self._browser_update_preview = _update_preview

        def _refresh_snapshot_list():
            names = [_norm(p.get("name", "")) for p in profiles if _norm(p.get("name", ""))]
            combo_snapshot["values"] = names
            # 同步 exe/serdir 的历史下拉
            exe_hist = sorted({ _norm(p.get("exe_path", "")) for p in profiles if _norm(p.get("exe_path", "")) })
            combo_exe["values"] = exe_hist
            ser_hist = sorted({ _norm(p.get("user_dir", p.get("user_data_dir", ""))) for p in profiles if _norm(p.get("user_dir", p.get("user_data_dir", ""))) })
            combo_userdir["values"] = ser_hist
            crx_hist = sorted({ _norm(p.get("crx_path", "")) for p in profiles if _norm(p.get("crx_path", "")) })
            if os.path.isfile(DEFAULT_CRX_PATH):
                crx_hist = sorted(set(crx_hist + [DEFAULT_CRX_PATH]))
            combo_crx["values"] = crx_hist

        def _on_snapshot_change(_event=None):
            name = _norm(var_snapshot.get())
            if not name:
                return
            p = next((p for p in profiles if _norm(p.get("name", "")) == name), None)
            if p:
                _apply_profile(p)
        # 暴露：按名称应用环境
        def _apply_profile_by_name(name: str) -> bool:
            nm = _norm(name)
            if not nm:
                return False
            p = next((p for p in profiles if _norm(p.get("name", "")) == nm), None)
            if not p:
                return False
            _apply_profile(p)
            return True

        self._browser_apply_profile = _apply_profile
        self._browser_apply_profile_by_name = _apply_profile_by_name

        combo_snapshot.bind("<<ComboboxSelected>>", _on_snapshot_change)
        # 支持“手动输入名称后回车/失焦”也触发加载（避免你说的“加载没有刷新”）
        combo_snapshot.bind("<Return>", _on_snapshot_change)
        combo_snapshot.bind("<FocusOut>", _on_snapshot_change)
        try:
            var_snapshot.trace_add("write", lambda *_a: _on_snapshot_change())
        except Exception:
            pass

        def _save_snapshot():
            name = _norm(var_snapshot.get())
            if not name:
                messagebox.showwarning("提示", "请先填写/选择快照名称。", parent=self.root)
                return
            p = _profile_from_form(name=name)
            if not p["exe_path"]:
                messagebox.showwarning("提示", "请先填写 exe 路径。", parent=self.root)
                return
            idx = next((i for i, x in enumerate(profiles) if _norm(x.get("name", "")) == name), None)
            if idx is None:
                profiles.append(p)
            else:
                profiles[idx] = p
            save_saved_browser_profiles(profiles)
            self._log("已保存快照: " + name)
            try:
                messagebox.showinfo("保存成功", f"已保存快照：{name}", parent=self.root)
            except Exception:
                pass
            _refresh_snapshot_list()

        def _launch():
            nm = _norm(var_snapshot.get())
            p = next((p for p in profiles if _norm(p.get("name", "")) == nm), None)
            p = p if p else _profile_from_form()
            args = _build_args({
                "name": _norm(p.get("name", "")) or nm,
                "exe_path": _norm(p.get("exe_path", p.get("browser_path", var_exe.get()))),
                "fp": _norm(p.get("fp", var_fp.get())),
                "fp_font": _norm(p.get("fp_font", var_fp_font.get())),
                "language": _norm(p.get("language", var_lang.get())),
                "core": _norm(p.get("core", var_core.get())),
                "vendor": _norm(p.get("vendor", var_vendor.get())),
                "renderer": _norm(p.get("renderer", var_renderer.get())),
                "ua": _norm(p.get("ua", var_ua.get())),
                "win_version": _norm(p.get("win_version", var_winver.get())),
                "chrome_version": _norm(p.get("chrome_version", var_chromever.get())),
                "webrtc": _norm(p.get("webrtc", var_webrtc.get())),
                "timezone": _norm(p.get("timezone", var_timezone.get())),
                # 兜底：不允许为空，否则 _build_args 不会拼 user-data-dir
                "user_dir": _norm(p.get("user_dir", var_userdir.get())) or DEFAULT_USERDIR,
                # 关键：启动时也要把 crx_path 传进去，否则只预览有插件参数
                "crx_path": _norm(p.get("crx_path", var_crx.get())),
                "other": _norm(p.get("other", txt_other.get("1.0", tk.END))),
            })
            if not args:
                messagebox.showwarning("提示", "请先填写 exe 路径。", parent=self.root)
                return
            try:
                p = subprocess.Popen(args, cwd=os.path.dirname(args[0]) or None)
                # 记录最近一次启动的浏览器进程，供首页“一键关闭”使用
                try:
                    self._browser_proc = p
                    self._browser_last_pid = int(getattr(p, "pid", 0) or 0) or None
                except Exception:
                    pass
                self._log("已启动: " + (nm or "当前配置"))
            except Exception as e:
                messagebox.showerror("启动失败", str(e), parent=self.root)
        # 暴露启动函数（首页一键启动会调用）
        self._browser_launch = _launch

        def _close():
            try:
                self.nav.select(0)  # 返回“收件箱”
            except Exception:
                pass

        ttk.Button(btns, text="生成", command=_update_preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btns, text="清理", command=_clear_form).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btns, text="启动", command=_launch).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btns, text="关闭", command=_close).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btns, text="保存", command=_save_snapshot).pack(side=tk.LEFT, padx=(0, 8))

        _sync_renderer_list()
        _sync_chrome_versions_for_kernel()
        _sync_ua_from_kernel_or_chrome()
        _refresh_snapshot_list()
        _update_preview()

    def _remove_account(self):
        sel = (self.var_email.get() or "").strip()
        if not sel:
            messagebox.showwarning("未选择", "请先输入或选择要移除的邮箱。")
            return
        n_before = len(self._accounts)
        self._accounts = [a for a in self._accounts if (a.get("email") or "").strip() != sel]
        if len(self._accounts) >= n_before:
            messagebox.showinfo("提示", "该邮箱未在保存列表中。")
            return
        save_saved_accounts(self._accounts)
        self._refresh_account_combo()
        self.var_email.set("")
        self.var_pwd.set("")
        self.var_imap.set("")
        self._log("已移除: " + sel)

    def _on_import_accounts(self):
        """从 JSON 或 Excel 文件导入邮箱账号，并入已保存列表。"""
        path = filedialog.askopenfilename(
            title="选择账号文件 (JSON 或 Excel)",
            filetypes=[
                ("JSON / Excel", "*.json;*.xlsx"),
                ("JSON", "*.json"),
                ("Excel", "*.xlsx"),
                ("全部", "*.*"),
            ],
        )
        if not path or not os.path.isfile(path):
            return
        ext = os.path.splitext(path)[1].lower()
        items = []
        if ext == ".json":
            try:
                with open(path, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                arr = raw if isinstance(raw, list) else [raw]
                for o in arr:
                    if not isinstance(o, dict):
                        continue
                    email = (o.get("email") or "").strip()
                    pwd = (o.get("email_password") or o.get("password") or "").strip()
                    imap = (o.get("imap") or "").strip()
                    if email and pwd:
                        items.append({"email": email, "password": pwd, "imap": imap})
            except Exception as e:
                messagebox.showerror("导入失败", f"JSON 解析错误：{e}")
                return
        elif ext == ".xlsx":
            if not _HAS_OPENPYXL:
                messagebox.showerror("导入失败", "需安装 openpyxl 才能导入 Excel，请运行: pip install openpyxl")
                return
            try:
                from openpyxl import load_workbook

                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                if not rows:
                    messagebox.showinfo("导入", "Excel 文件为空。")
                    return
                # 首行作表头，不区分大小写查找列
                hdr = [str((r or "")).strip() for r in rows[0]]
                hdr_lower = [s.lower() for s in hdr]
                email_col = _find_col(hdr_lower, ["email", "e-mail", "邮箱", "mail"])
                pwd_col = _find_col(hdr_lower, ["email_password", "password", "密码", "pwd"])
                imap_col = _find_col(hdr_lower, ["imap", "imap服务器", "imap_server"])
                if email_col is None:
                    email_col = 0
                if pwd_col is None:
                    pwd_col = 1
                for row in rows[1:]:
                    row = list(row) if row else []
                    email = str(row[email_col] or "").strip() if email_col < len(row) else ""
                    pwd = str(row[pwd_col] or "").strip() if pwd_col < len(row) else ""
                    imap = (str(row[imap_col] or "").strip() if imap_col is not None and imap_col < len(row) else "") or ""
                    if email and pwd:
                        items.append({"email": email, "password": pwd, "imap": imap})
            except Exception as e:
                messagebox.showerror("导入失败", f"Excel 解析错误：{e}")
                return
        else:
            messagebox.showwarning("导入", "仅支持 .json 和 .xlsx 文件。")
            return
        if not items:
            messagebox.showinfo("导入", "未解析到有效账号（需包含邮箱和密码）。")
            return
        # 合并到已保存：同邮箱则更新，否则追加
        exist_emails = {(a.get("email") or "").strip().lower() for a in self._accounts}
        added, updated = 0, 0
        for it in items:
            em = (it.get("email") or "").strip()
            em_lower = em.lower()
            if em_lower in exist_emails:
                for a in self._accounts:
                    if (a.get("email") or "").strip().lower() == em_lower:
                        a["password"] = it.get("password", "")
                        a["imap"] = it.get("imap", "")
                        updated += 1
                        break
            else:
                self._accounts.append(it)
                exist_emails.add(em_lower)
                added += 1
        save_saved_accounts(self._accounts)
        self._refresh_account_combo()
        self._log("导入完成：新增 %d，更新 %d（共 %d 条）" % (added, updated, len(items)))
        messagebox.showinfo("导入完成", "新增 %d 个，更新 %d 个，共处理 %d 条。" % (added, updated, len(items)))

    def _copy_code(self):
        s = (self.var_code.get() or "").strip()
        if s:
            self.root.clipboard_clear()
            self.root.clipboard_append(s)
            self._log("验证码已复制: " + s)
        else:
            messagebox.showinfo("提示", "暂无验证码可复制。")

    def _toggle_pwd_visible(self):
        """切换密码输入框显示/隐藏"""
        self._pwd_visible = not self._pwd_visible
        if hasattr(self, "entry_pwd") and self.entry_pwd:
            self.entry_pwd.config(show="" if self._pwd_visible else "•")
        if hasattr(self, "icon_eye") and self.icon_eye:
            # 可见：眼睛无斜杠；隐藏：眼睛带斜杠
            self.icon_eye.set_eye_open(self._pwd_visible)
        try:
            if hasattr(self, "entry_pwd") and self.entry_pwd:
                self.entry_pwd.focus_set()
        except Exception:
            pass

    def _copy_pwd(self):
        """复制密码到剪贴板"""
        s = (self.var_pwd.get() or "")
        if not s.strip():
            messagebox.showinfo("提示", "密码为空，无法复制。")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(s)
        self._log("密码已复制到剪贴板")
        try:
            if hasattr(self, "entry_pwd") and self.entry_pwd:
                self.entry_pwd.focus_set()
        except Exception:
            pass

    def _show_emails(self, items: list[dict]):
        def _do():
            first_code = ""
            self.txt_inbox.delete("1.0", tk.END)
            if not items:
                self.txt_inbox.insert(tk.END, "（无邮件或读取为空）\n")
                self.var_code.set("")
                return
            sep = "─" * 50
            for i, m in enumerate(items, 1):
                # 卡片分隔（非首封时先加空行）
                if i > 1:
                    self.txt_inbox.insert(tk.END, "\n")
                self.txt_inbox.insert(tk.END, sep + "\n", "sep")
                # 标题（如 Verification Code）
                subj = m.get("subject", "") or "(无主题)"
                self.txt_inbox.insert(tk.END, subj + "\n", "preview_title")
                self.txt_inbox.insert(tk.END, "\n")
                # 验证码：突出显示
                code = m.get("code", "") or ""
                if code:
                    self.txt_inbox.insert(tk.END, "  验证码: " + code + "\n\n", "code_block")
                    if not first_code:
                        first_code = code
                # 正文（纯文本展示，非 HTML 源码）
                prev = m.get("body_display", "") or ""
                if prev:
                    self.txt_inbox.insert(tk.END, prev + "\n\n", "body")
                # 发件人、日期（底部元信息）
                self.txt_inbox.insert(tk.END, "  发件人: " + (m.get("from_addr", "") or "") + "\n", "meta")
                self.txt_inbox.insert(tk.END, "  日期:   " + (m.get("date", "") or "") + "\n", "meta")
            self.var_code.set(first_code)
            self.txt_inbox.see("1.0")
        self.root.after(0, _do)

    def _build_ui(self):
        # 顶部导航栏：用标签页做“工具切换”
        self.nav = ttk.Notebook(self.root)
        self.nav.pack(fill=tk.BOTH, expand=True)

        tab_home = ttk.Frame(self.nav, padding=12)
        tab_inbox = ttk.Frame(self.nav, padding=12)
        tab_browser = ttk.Frame(self.nav, padding=12)
        tab_accounts = ttk.Frame(self.nav, padding=12)
        self.nav.add(tab_home, text="首页")
        self.nav.add(tab_inbox, text="收件箱")
        self.nav.add(tab_browser, text="浏览器")
        self.nav.add(tab_accounts, text="账号管理")

        main = tab_inbox

        # 圆角输入框的配色（尽量跟随主题）
        style = ttk.Style(self.root)
        field_bg = _style_lookup(style, "TEntry", "fieldbackground", "#ffffff")
        outer_bg = _style_lookup(style, "TFrame", "background", self.root.cget("bg"))
        fg = _style_lookup(style, "TEntry", "foreground", "#111827")
        try:
            style.configure("Flat.TCombobox", borderwidth=0, relief="flat", padding=0)
        except Exception:
            pass

        # --- 代理 (GOST) 放上边 ---
        f_proxy = ttk.LabelFrame(main, text="代理 (GOST)", padding=8)
        f_proxy.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(f_proxy, text="转发地址:").grid(row=0, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_forward = tk.StringVar(value="socks5://127.0.0.1:10809")
        self.box_forward = RoundedBox(f_proxy, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=4, pad_y=1)
        self.box_forward.grid(row=0, column=1, sticky=tk.EW, padx=2, pady=4)
        self.entry_forward = tk.Entry(
            self.box_forward.inner,
            textvariable=self.var_forward,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.entry_forward.pack(fill=tk.BOTH, expand=True)
        self.entry_forward.bind("<FocusIn>", lambda _e: self.box_forward.set_focused(True))
        self.entry_forward.bind("<FocusOut>", lambda _e: self.box_forward.set_focused(False))
        ttk.Label(f_proxy, text="（默认 10809，可改）").grid(row=0, column=2, sticky=tk.W, padx=4)

        ttk.Label(f_proxy, text="原始代理:").grid(row=1, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_share_url = tk.StringVar()
        # 下拉框：轻微圆角，但内边距几乎为 0
        self.box_combo_proxy = RoundedBox(f_proxy, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=1, pad_y=1, radius=8)
        self.box_combo_proxy.grid(row=1, column=1, sticky=tk.EW, padx=2, pady=4)
        self.combo_proxy = ttk.Combobox(self.box_combo_proxy.inner, textvariable=self.var_share_url, width=42, style="Flat.TCombobox")
        self.combo_proxy.pack(fill=tk.BOTH, expand=True)
        self.combo_proxy.bind("<FocusIn>", lambda _e: self.box_combo_proxy.set_focused(True))
        self.combo_proxy.bind("<FocusOut>", lambda _e: self.box_combo_proxy.set_focused(False))
        self.combo_proxy.bind("<<ComboboxSelected>>", self._on_proxy_select)
        ttk.Label(f_proxy, text="（下拉选已保存或直接输入）").grid(row=1, column=2, columnspan=2, sticky=tk.W, padx=6)

        ttk.Label(f_proxy, text="本地端口:").grid(row=2, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_proxy_port = tk.StringVar(value="50681")
        self.box_proxy_port = RoundedBox(f_proxy, fill=field_bg, outer_bg=outer_bg, height=30, width=140, pad_x=4, pad_y=1)
        self.box_proxy_port.grid(row=2, column=1, sticky=tk.W, padx=2, pady=4)
        self.entry_proxy_port = tk.Entry(
            self.box_proxy_port.inner,
            textvariable=self.var_proxy_port,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.entry_proxy_port.pack(fill=tk.BOTH, expand=True)
        self.entry_proxy_port.bind("<FocusIn>", lambda _e: self.box_proxy_port.set_focused(True))
        self.entry_proxy_port.bind("<FocusOut>", lambda _e: self.box_proxy_port.set_focused(False))
        ttk.Label(f_proxy, text="GOST 监听端口").grid(row=2, column=2, sticky=tk.W, padx=4)

        f_px_btn = ttk.Frame(f_proxy)
        f_px_btn.grid(row=3, column=1, sticky=tk.W, padx=2, pady=4)
        ttk.Button(f_px_btn, text="保存", command=self._save_current_proxy).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(f_px_btn, text="移除", command=self._remove_proxy).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(f_px_btn, text="开启代理", command=self._on_start_proxy).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(f_px_btn, text="断开代理", command=self._on_stop_proxy).pack(side=tk.LEFT, padx=4)
        self.btn_test_proxy = ttk.Button(f_px_btn, text="测试代理", command=self._on_test_proxy)
        self.btn_test_proxy.pack(side=tk.LEFT, padx=4)
        self.lbl_proxy_status = ttk.Label(f_proxy, text="未开启")
        self.lbl_proxy_status.grid(row=3, column=2, sticky=tk.W, padx=4, pady=4)
        ttk.Label(f_proxy, text="代理测试:").grid(row=4, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        # 代理测试结果：可鼠标拖拽选择后复制（只读）
        self.var_proxy_test = tk.StringVar(value="—")
        self.box_proxy_test = RoundedBox(f_proxy, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=4, pad_y=1, radius=10)
        self.box_proxy_test.grid(row=4, column=1, columnspan=2, sticky=tk.EW, padx=2, pady=4)
        self.entry_proxy_test = tk.Entry(
            self.box_proxy_test.inner,
            textvariable=self.var_proxy_test,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg="#666",
            insertbackground=fg,
            readonlybackground=field_bg,
        )
        self.entry_proxy_test.pack(fill=tk.BOTH, expand=True)
        self.entry_proxy_test.configure(state="readonly")
        self.entry_proxy_test.bind("<FocusIn>", lambda _e: self.box_proxy_test.set_focused(True))
        self.entry_proxy_test.bind("<FocusOut>", lambda _e: self.box_proxy_test.set_focused(False))
        f_proxy.columnconfigure(1, weight=1)

        # --- 登录区 放下边 ---
        f = ttk.LabelFrame(main, text="登录", padding=8)
        f.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(f, text="邮箱:").grid(row=0, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_email = tk.StringVar()
        # 下拉框：轻微圆角，但内边距几乎为 0
        self.box_combo_email = RoundedBox(f, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=1, pad_y=1, radius=8)
        self.box_combo_email.grid(row=0, column=1, sticky=tk.EW, padx=2, pady=4)
        self.combo_email = ttk.Combobox(self.box_combo_email.inner, textvariable=self.var_email, width=42, style="Flat.TCombobox")
        self.combo_email.pack(fill=tk.BOTH, expand=True)
        self.combo_email.bind("<FocusIn>", lambda _e: self.box_combo_email.set_focused(True))
        self.combo_email.bind("<FocusOut>", lambda _e: self.box_combo_email.set_focused(False))
        self.combo_email.bind("<<ComboboxSelected>>", self._on_email_select)
        self.combo_email.bind("<KeyRelease>", self._on_email_keyrelease)
        self.combo_email.bind("<Return>", self._on_email_commit)
        self.combo_email.bind("<FocusOut>", self._on_email_commit)
        self.var_email.trace_add("write", self._on_email_var_changed)
        ttk.Label(f, text="（下拉选已保存或直接输入）").grid(row=0, column=2, columnspan=2, sticky=tk.W, padx=6)

        ttk.Label(f, text="密码:").grid(row=1, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_pwd = tk.StringVar()
        self.box_pwd = RoundedBox(f, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=4, pad_y=1)
        self.box_pwd.grid(row=1, column=1, sticky=tk.EW, padx=2, pady=4)
        pwd_inner = tk.Frame(self.box_pwd.inner, bg=field_bg)
        pwd_inner.pack(fill=tk.BOTH, expand=True)

        # 右侧两个小图标（内部）：复制 / 眼睛
        icons = tk.Frame(pwd_inner, bg=field_bg)
        icons.pack(side=tk.RIGHT, padx=(2, 2))
        self.icon_copy_pwd = IconButton(icons, kind="copy", command=self._copy_pwd, bg=field_bg)
        self.icon_copy_pwd.pack(side=tk.RIGHT, padx=(2, 0))
        self.icon_eye = IconButton(icons, kind="eye", command=self._toggle_pwd_visible, bg=field_bg)
        # 初始为隐藏：眼睛带斜杠
        self.icon_eye.set_eye_open(False)
        self.icon_eye.pack(side=tk.RIGHT, padx=(0, 0))

        self.entry_pwd = tk.Entry(
            pwd_inner,
            textvariable=self.var_pwd,
            show="•",
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.entry_pwd.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.entry_pwd.bind("<FocusIn>", lambda _e: self.box_pwd.set_focused(True))
        self.entry_pwd.bind("<FocusOut>", lambda _e: self.box_pwd.set_focused(False))

        ttk.Label(f, text="IMAP 服务器:").grid(row=2, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        self.var_imap = tk.StringVar(value="")
        self.box_imap = RoundedBox(f, fill=field_bg, outer_bg=outer_bg, height=30, pad_x=4, pad_y=1)
        self.box_imap.grid(row=2, column=1, sticky=tk.EW, padx=2, pady=4)
        self.entry_imap = tk.Entry(
            self.box_imap.inner,
            textvariable=self.var_imap,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.entry_imap.pack(fill=tk.BOTH, expand=True)
        self.entry_imap.bind("<FocusIn>", lambda _e: self.box_imap.set_focused(True))
        self.entry_imap.bind("<FocusOut>", lambda _e: self.box_imap.set_focused(False))
        ttk.Label(f, text="（留空则按域名自动推断）").grid(row=2, column=2, columnspan=2, sticky=tk.W, padx=6)

        f_btns = ttk.Frame(f)
        f_btns.grid(row=3, column=1, sticky=tk.W, padx=2, pady=4)
        ttk.Button(f_btns, text="保存", command=self._save_current_account).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(f_btns, text="移除", command=self._remove_account).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(f_btns, text="导入", command=self._on_import_accounts).pack(side=tk.LEFT)

        f.columnconfigure(1, weight=1)

        # 账号管理页（原弹窗改为主窗口内页）
        # 注意：需要先创建 self.var_email/self.var_pwd/self.var_share_url 才能用于默认值
        self._build_fb_account_manager_tab(tab_accounts)

        # 浏览器页：放在这里构建
        self._build_browser_settings_tab(tab_browser)
        # 浏览器页创建完成后：让账号管理的“浏览器环境”默认展示当前选择
        try:
            cb = getattr(self, "_accounts_set_browser_profile_default", None)
            if callable(cb):
                cb(force=True)
        except Exception:
            pass

        # 首页（最后构建：需要上面变量/函数都就绪，便于一键启动联动）
        self._build_home_tab(tab_home)

        # --- 按钮与验证码 ---
        f_btn = ttk.Frame(main)
        f_btn.pack(fill=tk.X, pady=6)

        self.btn_login = ttk.Button(f_btn, text="登录并读取最近 5 封", command=self._on_login)
        self.btn_login.pack(side=tk.LEFT, padx=(0, 8))
        self.lbl_status = ttk.Label(f_btn, text="就绪")
        self.lbl_status.pack(side=tk.LEFT, padx=8)
        ttk.Label(f_btn, text="  验证码:").pack(side=tk.LEFT, padx=(16, 4))
        self.var_code = tk.StringVar(value="")
        self.box_code = RoundedBox(f_btn, fill=field_bg, outer_bg=outer_bg, height=28, width=140, pad_x=4, pad_y=1)
        self.box_code.pack(side=tk.LEFT, padx=2)
        self.entry_code = tk.Entry(
            self.box_code.inner,
            textvariable=self.var_code,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.entry_code.pack(fill=tk.BOTH, expand=True)
        self.entry_code.bind("<FocusIn>", lambda _e: self.box_code.set_focused(True))
        self.entry_code.bind("<FocusOut>", lambda _e: self.box_code.set_focused(False))
        ttk.Button(f_btn, text="复制", command=self._copy_code).pack(side=tk.LEFT, padx=4)

        # --- 邮件预览（可滚动） ---
        f_inbox = ttk.LabelFrame(main, text="最近 5 封邮件 · 预览", padding=6)
        f_inbox.pack(fill=tk.BOTH, expand=True, pady=8)

        self.box_inbox = RoundedBox(f_inbox, fill=field_bg, outer_bg=outer_bg)
        self.box_inbox.pack(fill=tk.BOTH, expand=True)
        self.txt_inbox = scrolledtext.ScrolledText(
            self.box_inbox.inner,
            height=18,
            font=("Segoe UI", 10),
            wrap=tk.WORD,
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.txt_inbox.pack(fill=tk.BOTH, expand=True)
        self.txt_inbox.bind("<FocusIn>", lambda _e: self.box_inbox.set_focused(True))
        self.txt_inbox.bind("<FocusOut>", lambda _e: self.box_inbox.set_focused(False))
        self.txt_inbox.tag_config("sep", foreground="#ccc")
        self.txt_inbox.tag_config("preview_title", font=("Segoe UI", 12, "bold"))
        self.txt_inbox.tag_config("code_block", font=("Segoe UI", 11, "bold"), foreground="#0066cc")
        self.txt_inbox.tag_config("body", font=("Segoe UI", 10))
        self.txt_inbox.tag_config("meta", font=("Segoe UI", 9), foreground="#666")

        # --- 日志 ---
        f_log = ttk.LabelFrame(main, text="日志", padding=4)
        f_log.pack(fill=tk.X, pady=(0, 4))

        # 日志更矮一些，给邮件预览更多空间
        self.box_log = RoundedBox(f_log, fill=field_bg, outer_bg=outer_bg, height=64)
        self.box_log.pack(fill=tk.X)
        self.log = scrolledtext.ScrolledText(
            self.box_log.inner,
            height=3,
            font=("Consolas", 9),
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=field_bg,
            fg=fg,
            insertbackground=fg,
        )
        self.log.pack(fill=tk.BOTH, expand=True)
        self.log.bind("<FocusIn>", lambda _e: self.box_log.set_focused(True))
        self.log.bind("<FocusOut>", lambda _e: self.box_log.set_focused(False))

    def _on_start_proxy(self):
        """开启 GOST 代理：-L=socks5://127.0.0.1:{port} -F={forward} -F={share_url}"""
        forward = (self.var_forward.get() or "").strip()
        share_url = (self.var_share_url.get() or "").strip()
        port_s = (self.var_proxy_port.get() or "").strip()
        if not share_url:
            messagebox.showwarning("输入有误", "请填写原始代理 (share_url)。")
            return
        if not port_s:
            messagebox.showwarning("输入有误", "请指定本地端口。")
            return
        try:
            port = int(port_s)
            if port < 1 or port > 65535:
                raise ValueError("端口越界")
        except ValueError:
            messagebox.showwarning("输入有误", "本地端口须为 1～65535 的整数。")
            return
        if not forward:
            forward = "socks5://127.0.0.1:10809"
        gost_path = os.path.join(_app_dir(), "gost.exe")
        if not os.path.isfile(gost_path):
            messagebox.showerror("未找到 GOST", f"请将 gost.exe 放在程序同目录：\n{gost_path}")
            return
        if self._gost_proc is not None and self._gost_proc.poll() is None:
            messagebox.showinfo("提示", "代理已在运行，可先断开再重新开启。")
            return
        cmd = [
            gost_path,
            "-L=socks5://127.0.0.1:" + str(port),
            "-F=" + forward,
            "-F=" + share_url,
        ]
        try:
            self._gost_proc = subprocess.Popen(
                cmd,
                creationflags=0x08000000 if sys.platform == "win32" else 0,  # CREATE_NO_WINDOW
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            self.lbl_proxy_status.config(text=f"已开启 127.0.0.1:{port}")
            self._log(f"代理已开启: 127.0.0.1:{port} <- {forward} <- {share_url}")
        except Exception as e:
            messagebox.showerror("启动失败", str(e))
            self._log("代理启动失败: " + str(e))

    def _on_stop_proxy(self):
        """断开 GOST 代理进程"""
        if self._gost_proc is None:
            self.lbl_proxy_status.config(text="未开启")
            return
        try:
            self._gost_proc.terminate()
            self._gost_proc.wait(timeout=5)
        except Exception:
            try:
                self._gost_proc.kill()
            except Exception:
                pass
        self._gost_proc = None
        self.lbl_proxy_status.config(text="未开启")
        self._log("代理已断开")

    def _on_test_proxy(self):
        """通过代理请求 https://ipinfo.io/json，展示 IP、国家地区、时区（urllib + PySocks）"""
        if not _HAS_PYSOCKS:
            messagebox.showwarning("缺少依赖", "请安装 PySocks: pip install PySocks")
            return
        port_s = (self.var_proxy_port.get() or "").strip()
        if not port_s:
            messagebox.showwarning("输入有误", "请先填写本地端口。")
            return
        try:
            port = int(port_s)
            if port < 1 or port > 65535:
                raise ValueError("端口越界")
        except ValueError:
            messagebox.showwarning("输入有误", "本地端口须为 1～65535 的整数。")
            return
        self.var_proxy_test.set("测试中…")
        try:
            self.entry_proxy_test.config(fg="#666")
        except Exception:
            pass
        self.btn_test_proxy.config(state=tk.DISABLED)

        def _work():
            err_msg = ""
            try: 
                proxies = {"http": "socks5://127.0.0.1:" + str(port), "https": "socks5://127.0.0.1:" + str(port)}
                print(proxies)
                response = requests.get(PROXY_TEST_URL, proxies=proxies, timeout=60)
                d = response.json()
                print(d)
                ip = d.get("ip", "")
                country = d.get("country", "")
                region = d.get("region", "") or d.get("city", "")
                timezone = d.get("timezone", "")
                area = " / ".join(x for x in (country, region) if x)
                line = f"IP: {ip} | 国家/地区: {area} | 时区: {timezone}"
                self._log("[代理测试] " + line)

                def _ui():
                    self.var_proxy_test.set(line)
                    try:
                        self.entry_proxy_test.config(fg="#0a0")
                    except Exception:
                        pass
                    # 联动回填到“浏览器”页：webrtc=ip，timezone=timezone
                    try:
                        self._last_proxy_test_ip = ip
                        self._last_proxy_test_timezone = timezone
                        if hasattr(self, "var_browser_webrtc") and self.var_browser_webrtc is not None:
                            self.var_browser_webrtc.set(ip or "")
                        if hasattr(self, "var_browser_timezone") and self.var_browser_timezone is not None:
                            self.var_browser_timezone.set(timezone or "")
                    except Exception:
                        pass
                    self.btn_test_proxy.config(state=tk.NORMAL)

                self.root.after(0, _ui)
            except OSError as e:
                err_msg = f"代理连接失败: {e}"
            except urllib.error.URLError as e:
                err_msg = f"请求失败: {e.reason}"
            except urllib.error.HTTPError as e:
                err_msg = f"HTTP {e.code}: {e.reason}"
            except Exception as e:
                err_msg = str(e)
            if err_msg:
                self._log("[代理测试] " + err_msg)

                def _ui_err():
                    self.var_proxy_test.set(err_msg)
                    try:
                        self.entry_proxy_test.config(fg="#c00")
                    except Exception:
                        pass
                    self.btn_test_proxy.config(state=tk.NORMAL)

                self.root.after(0, _ui_err)

        threading.Thread(target=_work, daemon=True).start()

    def _on_login(self):
        email_addr = (self.var_email.get() or "").strip()
        pwd = (self.var_pwd.get() or "").strip()
        imap_custom = (self.var_imap.get() or "").strip()

        if not email_addr or not pwd:
            messagebox.showwarning("输入有误", "请填写邮箱和密码。")
            return

        host, port = guess_imap_server(email_addr)
        if imap_custom:
            if ":" in imap_custom:
                host, p = imap_custom.rsplit(":", 1)
                try:
                    port = int(p)
                except ValueError:
                    port = 993
            else:
                host = imap_custom
                port = 993

        self._set_loading(True)
        self._log(f"正在连接 {host}:{port} …")

        def _work():
            try:
                result = fetch_recent_emails(email_addr, pwd, limit=5, imap_host=host, imap_port=port)
                if isinstance(result, str):
                    self._log(result)
                    self.root.after(0, lambda: messagebox.showerror("读取失败", result))
                    self._show_emails([])
                else:
                    self._log(f"成功读取 {len(result)} 封邮件")
                    self._show_emails(result)
                    self.root.after(0, self._save_current_account)  # 登录成功则保存/更新账号
            finally:
                self._set_loading(False)

        t = threading.Thread(target=_work, daemon=True)
        t.start()


def main():
    app = EmailInboxReaderApp()
    app.root.mainloop()


if __name__ == "__main__":
    main()
